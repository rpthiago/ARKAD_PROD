"""
coleta_layhome_aovivo.py — coleta ao vivo de odds in-play do Lay Home
======================================================================
A cada dia, adiciona ao ledger os jogos que o Lay Home sinalizaria (mandante
favorito, odd 1.4-2.5, ligas principais) com colunas EM BRANCO para você anotar,
DO SEU SITE, a odd de vitória da casa na ABERTURA (min 0), min 30 e min 45.
O Placar_final e o Momento_gols se preenchem sozinhos quando o jogo termina.

Objetivo: acumular, com odd de MESMA FONTE (entrada+saídas do seu site), jogos
suficientes de "casa venceu" para confirmar (ou derrubar) o achado de que o
STOP-LOSS vira o Lay Home positivo (+~8% no estudo, mas IC ainda cruzava zero).

Uso:  python coleta_layhome_aovivo.py [YYYY-MM-DD]
"""
import os, sys, datetime, unicodedata, re, ast
import pandas as pd, numpy as np
sys.stdout.reconfigure(encoding="utf-8")

LEDGER = "coleta_layhome_aovivo.xlsx"
MAJORS = ["ENGLAND 1","SPAIN 1","ITALY 1","GERMANY 1","FRANCE 1","BRAZIL 1",
          "PORTUGAL 1","NETHERLANDS 1","ENGLAND 2","BRAZIL 2","ARGENTINA 1"]
ODD_MIN, ODD_MAX = 1.40, 2.50
COLS = ["Date","Horario","Liga","Mandante","Visitante","Odd_H_ref_b365",
        "PREENCHER_odd_abertura","PREENCHER_odd_min30","PREENCHER_odd_min45",
        "Placar_final","Momento_gols","status","obs"]

def _canon(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii","ignore").decode().lower()
    return re.sub(r"[^a-z0-9]", "", s)

def carregar():
    if os.path.exists(LEDGER) and os.path.getsize(LEDGER) > 0:
        try: return pd.read_excel(LEDGER, sheet_name="Jogos", dtype=str)
        except Exception: pass
    return pd.DataFrame(columns=COLS)

def salvar_xlsx(df):
    """Escreve o ledger em .xlsx formatado (colunas a preencher em amarelo),
    preservando o que o usuario ja digitou. Retry se o Excel estiver aberto."""
    import time
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    fillc = {"PREENCHER_odd_abertura","PREENCHER_odd_min30","PREENCHER_odd_min45","obs"}
    widths = {"Date":11,"Horario":8,"Liga":13,"Mandante":18,"Visitante":18,"Odd_H_ref_b365":14,
              "PREENCHER_odd_abertura":16,"PREENCHER_odd_min30":14,"PREENCHER_odd_min45":14,
              "Placar_final":11,"Momento_gols":26,"status":10,"obs":14}
    for tent in range(6):
        try:
            wb=Workbook(); ws=wb.active; ws.title="Jogos"
            fh=PatternFill("solid",fgColor="305496"); fy=PatternFill("solid",fgColor="FFF2CC"); fr=PatternFill("solid",fgColor="E2EFDA")
            white=Font(color="FFFFFF",bold=True); thin=Border(*[Side(style="thin",color="D0D0D0")]*4)
            for j,c in enumerate(COLS,1):
                cell=ws.cell(1,j,c); cell.font=white; cell.fill=fh
                cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=thin
            for i,(_,row) in enumerate(df.iterrows(),2):
                for j,c in enumerate(COLS,1):
                    v=row.get(c,""); cell=ws.cell(i,j,"" if pd.isna(v) else v); cell.border=thin
                    if c in fillc: cell.fill=fy
                    elif c=="Odd_H_ref_b365": cell.fill=fr
            for j,c in enumerate(COLS,1): ws.column_dimensions[get_column_letter(j)].width=widths.get(c,12)
            ws.freeze_panes="A2"; ws.row_dimensions[1].height=28
            wb.save(LEDGER); return True
        except PermissionError:
            if tent==5: return False
            time.sleep(2)
    return False

def sinais_do_dia(date_str):
    from b365_data_utils import fetch_b365_daily
    d = fetch_b365_daily(date_str)
    if d is None or d.empty: return []
    d = d.copy()
    d["oh"] = pd.to_numeric(d.get("Odd_H_FT"), errors="coerce")
    d["lg"] = d.get("League","").astype(str).str.upper().str.strip()
    d = d[(d["lg"].isin(MAJORS)) & (d["oh"] >= ODD_MIN) & (d["oh"] <= ODD_MAX)]
    out = []
    for _, r in d.iterrows():
        out.append(dict(Date=date_str, Horario=str(r.get("Time","") or ""), Liga=r.get("League",""),
                        Mandante=r.get("Home",""), Visitante=r.get("Away",""), Odd_H_ref_b365=round(float(r["oh"]),2),
                        PREENCHER_odd_abertura="", PREENCHER_odd_min30="", PREENCHER_odd_min45="",
                        Placar_final="", Momento_gols="", status="AGUARDA", obs=""))
    return out

def liquidar(led):
    """Preenche Placar_final e Momento_gols dos jogos ja disputados."""
    pend = led[led["status"] == "AGUARDA"]
    if pend.empty: return led, 0
    from b365_data_utils import load_b365_historical
    h = load_b365_historical().copy()
    h["d"] = pd.to_datetime(h["Date"], errors="coerce").dt.strftime("%Y-%m-%d")
    h["ch"] = h["Home"].map(_canon); h["ca"] = h["Away"].map(_canon)
    h["gh"] = pd.to_numeric(h.get("Goals_H_FT"), errors="coerce")
    h["ga"] = pd.to_numeric(h.get("Goals_A_FT"), errors="coerce")
    idx = {}
    for _, r in h.dropna(subset=["gh","ga"]).iterrows():
        idx[(r["d"], r["ch"], r["ca"])] = r
    def parse(v):
        try:
            s=str(v).strip(); return [int(str(x).split("+")[0]) for x in ast.literal_eval(s)] if s not in("","[]","nan") else []
        except: return []
    n = 0
    for i in pend.index:
        key = (str(led.at[i,"Date"])[:10], _canon(led.at[i,"Mandante"]), _canon(led.at[i,"Visitante"]))
        r = idx.get(key)
        if r is None: continue
        gh, ga = int(r["gh"]), int(r["ga"])
        mh, ma = parse(r.get("Goals_Min_H")), parse(r.get("Goals_Min_A"))
        mom = " | ".join([p for p in [
            "Casa:"+",".join(map(str,sorted(mh))) if mh else "",
            "Fora:"+",".join(map(str,sorted(ma))) if ma else ""] if p]) or "(sem gols)"
        led.at[i,"Placar_final"] = f"{gh}-{ga}"
        led.at[i,"Momento_gols"] = mom
        led.at[i,"status"] = "ENCERRADO"
        n += 1
    return led, n

def main():
    import pytz
    args = [a for a in sys.argv[1:] if re.match(r"\d{4}-\d{2}-\d{2}", a)]
    hoje = datetime.datetime.now(pytz.timezone("America/Sao_Paulo"))
    # sem data: processa HOJE + AMANHA (pra voce ja ver os jogos do dia seguinte)
    datas = args if args else [hoje.strftime("%Y-%m-%d"), (hoje + datetime.timedelta(days=1)).strftime("%Y-%m-%d")]
    led = carregar()
    led, nliq = liquidar(led)
    if nliq: print(f"  ✓ {nliq} jogo(s) encerrado(s) — placar preenchido.")
    ja = set(zip(led["Date"].astype(str), led["Mandante"].astype(str), led["Visitante"].astype(str))) if not led.empty else set()
    total_novos = 0
    for date_str in datas:
        novos = [s for s in sinais_do_dia(date_str) if (s["Date"], s["Mandante"], s["Visitante"]) not in ja]
        if novos:
            led = pd.concat([led, pd.DataFrame(novos)], ignore_index=True)
            ja |= {(s["Date"], s["Mandante"], s["Visitante"]) for s in novos}
            total_novos += len(novos)
            print(f"  ✓ {len(novos)} jogo(s) do Lay Home em {date_str}.")
    if not total_novos:
        print(f"  nenhum jogo novo do Lay Home em {datas} (fora de temporada / sem favoritos na faixa).")
    led = led.reindex(columns=COLS)
    if not salvar_xlsx(led):
        print("  ⚠️ planilha aberta no Excel — nao gravou. Feche o Excel e rode de novo (a lista abaixo ainda vale).")
    n_falta = ((led["status"]=="ENCERRADO") & (led["PREENCHER_odd_abertura"].fillna("")=="")).sum()
    print(f"  ledger: {len(led)} jogos | {(led['status']=='ENCERRADO').sum()} encerrados | {n_falta} encerrados ainda sem suas odds")

    # lista, na tela, os jogos que voce precisa acompanhar/preencher
    falta = led[led["PREENCHER_odd_abertura"].fillna("")==""].sort_values("Date")
    print("\n" + "="*64)
    print("  JOGOS LAY HOME PARA VOCE ANOTAR (abertura + min30 + min45):")
    print("="*64)
    if falta.empty:
        print("  (nenhum pendente — nada a fazer agora)")
    else:
        for _, r in falta.iterrows():
            tag = "a jogar" if r["status"]=="AGUARDA" else f"FIM {r['Placar_final']}"
            print(f"  {r['Date']} {str(r.get('Horario',''))[:5]:>5}  {str(r['Liga'])[:12]:<12}  {str(r['Mandante'])[:16]:<16} x {str(r['Visitante'])[:16]:<16}  (odd casa ~{r['Odd_H_ref_b365']}, {tag})")
    print(f"\n  Preencha no arquivo: {LEDGER}")

if __name__ == "__main__":
    main()
