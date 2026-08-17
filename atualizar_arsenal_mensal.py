"""
MOTOR DE ATUALIZAÇÃO AUTOMÁTICA MENSAL - ARSENAL COMPLETO PAPER TRADING
ARKAD_PROD

Este script atualiza automaticamente a planilha de acompanhamento do Arsenal Completo (01 a 31 de Agosto),
buscando os placares encerrados dos jogos futuros diretamente da API da Betfair, ESPN API e Bases Locais.

Como usar ao final do mês:
    python atualizar_arsenal_mensal.py
"""

import sys
import os
import shutil
import difflib
from pathlib import Path
import pandas as pd
import numpy as np
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from futpythontrader_client import _get_api_token, _build_headers

STAKE_UNIDADE = 100.0
COMISSAO_BETFAIR = 0.05

def norm_str(s: str) -> str:
    return ''.join(c for c in str(s).lower() if c.isalnum())

def carregar_base_placares_integrada() -> dict:
    """Busca placares encerrados de todas as fontes disponíveis (Bases Locais + ESPN API)."""
    db_scores = {}
    
    # 1. Leitura das Bases Locais em CSV
    sources_local = [
        'Bases_de_Dados_API_FutPythonTrader_Betfair_FRESH.csv',
        'Bases_de_Dados_API_FutPythonTrader_Bet365.csv',
        'Resultados_2026_Full.csv'
    ]
    
    for fpath in sources_local:
        p = Path(fpath)
        if p.exists():
            try:
                df = pd.read_csv(p, low_memory=False)
                d_col = [c for c in df.columns if 'date' in c.lower() or 'data' in c.lower()][0]
                h_col = 'Home' if 'Home' in df.columns else ('Home_Team' if 'Home_Team' in df.columns else ('Mandante' if 'Mandante' in df.columns else None))
                a_col = 'Away' if 'Away' in df.columns else ('Away_Team' if 'Away_Team' in df.columns else ('Visitante' if 'Visitante' in df.columns else None))
                gh_col = [c for c in df.columns if 'goals_h_ft' in c.lower() or 'gols mandante' in c.lower()][0]
                ga_col = [c for c in df.columns if 'goals_a_ft' in c.lower() or 'gols visitante' in c.lower()][0]
                
                df['d_str'] = pd.to_datetime(df[d_col], errors='coerce').dt.strftime('%Y-%m-%d')
                
                for _, r in df.iterrows():
                    dt = r['d_str']
                    if pd.notna(dt) and dt >= '2026-08-01':
                        gh = r[gh_col]
                        ga = r[ga_col]
                        if pd.notna(gh) and pd.notna(ga):
                            try:
                                gh_i = int(float(gh))
                                ga_i = int(float(ga))
                                hk = norm_str(r[h_col])
                                ak = norm_str(r[a_col])
                                db_scores[(dt, hk, ak)] = (gh_i, ga_i)
                                db_scores[(dt, hk[:5], ak[:5])] = (gh_i, ga_i)
                            except Exception:
                                pass
            except Exception as e:
                print(f"Aviso ao ler {fpath}: {e}")
                
    # 2. Leitura da ESPN API em tempo real para os dias do mês
    print("Consultando API ao vivo de placares da ESPN...")
    for day in range(1, 32):
        date_str = f"202608{day:02d}"
        url = f"https://site.api.espn.com/apis/site/v2/sports/soccer/all/scoreboard?dates={date_str}&limit=1000"
        try:
            r = requests.get(url, timeout=10)
            if r.status_code == 200:
                data = r.json()
                events = data.get('events', [])
                for ev in events:
                    try:
                        competitions = ev.get('competitions', [])[0]
                        competitors = competitions.get('competitors', [])
                        home_comp = [c for c in competitors if c.get('homeAway') == 'home'][0]
                        away_comp = [c for c in competitors if c.get('homeAway') == 'away'][0]
                        
                        h_name = home_comp.get('team', {}).get('displayName', '')
                        a_name = away_comp.get('team', {}).get('displayName', '')
                        
                        # Verifica se o jogo está finalizado
                        status_type = competitions.get('status', {}).get('type', {}).get('name', '')
                        if status_type in ['STATUS_FULL_TIME', 'STATUS_FINAL', 'STATUS_APPROVED', 'FULL_TIME']:
                            h_score = int(home_comp.get('score', 0))
                            a_score = int(away_comp.get('score', 0))
                            
                            dt_fmt = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
                            hk = norm_str(h_name)
                            ak = norm_str(a_name)
                            db_scores[(dt_fmt, hk, ak)] = (h_score, a_score)
                            db_scores[(dt_fmt, hk[:5], ak[:5])] = (h_score, a_score)
                    except Exception:
                        pass
        except Exception:
            pass
            
    return db_scores

def atualizar_planilha_arsenal():
    """Lê paper_trading_forward_setembro_2026.csv, cruza com placares e gera os Excels formatados."""
    csv_file = "paper_trading_forward_setembro_2026.csv"
    if not os.path.exists(csv_file):
        print(f"Erro: Arquivo {csv_file} não encontrado!")
        return

    df_pt = pd.read_csv(csv_file)
    print(f"Carregados {len(df_pt)} sinais do paper trading.")
    
    db_scores = carregar_base_placares_integrada()
    print(f"Base integrada de placares com {len(db_scores)} partidas finalizadas registradas.")
    
    out_rows = []
    greens_count = 0
    reds_count = 0
    pending_count = 0
    pnl_total = 0.0
    
    by_method = {}
    
    for idx, r in df_pt.iterrows():
        dt = str(r.get('data', ''))[:10]
        horario = str(r.get('horario', ''))[:5] if pd.notna(r.get('horario')) else ''
        liga = str(r.get('liga', ''))
        confronto = str(r.get('jogo', ''))
        metodo = str(r.get('metodo', '')).strip()
        lado = str(r.get('lado', '')).upper()
        odd = float(r.get('odd_execucao', 0.0)) if pd.notna(r.get('odd_execucao')) else 2.0
        
        if metodo not in by_method:
            by_method[metodo] = {'GREEN': 0, 'RED': 0, 'PENDING': 0, 'PNL': 0.0}
            
        gh, ga = None, None
        
        # 1. Verifica se já há placar no CSV original
        if 'Gols Mandante' in r and pd.notna(r['Gols Mandante']) and str(r['Gols Mandante']) != '' and 'Gols Visitante' in r and pd.notna(r['Gols Visitante']) and str(r['Gols Visitante']) != '':
            try:
                gh = int(float(r['Gols Mandante']))
                ga = int(float(r['Gols Visitante']))
            except Exception:
                pass
                
        # 2. Se não estiver no CSV, cruza com a base integrada
        if (gh is None or ga is None) and ' x ' in confronto:
            h_team, a_team = confronto.split(' x ')
            hk = norm_str(h_team)
            ak = norm_str(a_team)
            
            if (dt, hk, ak) in db_scores:
                gh, ga = db_scores[(dt, hk, ak)]
            elif (dt, hk[:5], ak[:5]) in db_scores:
                gh, ga = db_scores[(dt, hk[:5], ak[:5])]
            else:
                for (dt_k, hk_k, ak_k), sc in db_scores.items():
                    if dt_k == dt and len(hk_k) > 4:
                        sim_h = difflib.SequenceMatcher(None, hk, hk_k).ratio()
                        sim_a = difflib.SequenceMatcher(None, ak, ak_k).ratio()
                        if sim_h >= 0.45 and sim_a >= 0.45:
                            gh, ga = sc
                            break
                            
        res = ''
        pnl = 0.0
        
        if gh is not None and ga is not None:
            if 'Lay 0x3' in metodo:
                is_0x3 = (gh == 0 and ga == 3)
                win = not is_0x3
                res = 'GREEN' if win else 'RED'
                odd_capped = min(odd, 35.0)
                pnl = round(STAKE_UNIDADE * (1.0 - COMISSAO_BETFAIR), 2) if win else -round((odd_capped - 1.0) * STAKE_UNIDADE, 2)
            elif 'Lay 0x0' in metodo:
                is_0x0 = (gh == 0 and ga == 0)
                win = not is_0x0
                res = 'GREEN' if win else 'RED'
                odd_capped = min(odd, 12.0)
                pnl = round(STAKE_UNIDADE * (1.0 - COMISSAO_BETFAIR), 2) if win else -round((odd_capped - 1.0) * STAKE_UNIDADE, 2)
            elif 'Lay Draw' in metodo:
                is_draw = (gh == ga)
                win = not is_draw
                res = 'GREEN' if win else 'RED'
                pnl = round(STAKE_UNIDADE * (1.0 - COMISSAO_BETFAIR), 2) if win else -round((odd - 1.0) * STAKE_UNIDADE, 2)
            elif 'Over 2.5' in metodo:
                is_over25 = ((gh + ga) >= 3)
                win = is_over25
                res = 'GREEN' if win else 'RED'
                pnl = round((odd - 1.0) * STAKE_UNIDADE, 2) if win else -STAKE_UNIDADE
            elif 'BTTS Lay' in metodo:
                is_btts = (gh > 0 and ga > 0)
                win = not is_btts
                res = 'GREEN' if win else 'RED'
                pnl = round(STAKE_UNIDADE * (1.0 - COMISSAO_BETFAIR), 2) if win else -round((odd - 1.0) * STAKE_UNIDADE, 2)
            else:
                res = str(r.get('resultado', '')).upper()
                pnl = float(r.get('pnl_dolar', 0.0)) if pd.notna(r.get('pnl_dolar')) else 0.0
                win = (res == 'GREEN')
                
            if res == 'GREEN':
                greens_count += 1
                by_method[metodo]['GREEN'] += 1
            else:
                reds_count += 1
                by_method[metodo]['RED'] += 1
                
            pnl_total += pnl
            by_method[metodo]['PNL'] += pnl
        else:
            pending_count += 1
            by_method[metodo]['PENDING'] += 1
            
        out_rows.append({
            'Data': dt,
            'Horário': horario,
            'Liga': liga,
            'Confronto': confronto,
            'Método': metodo,
            'Lado': lado,
            'Odd Betfair': odd,
            'Gols Mandante': gh if gh is not None else '',
            'Gols Visitante': ga if ga is not None else '',
            'Resultado (GREEN/RED)': res,
            'Lucro/Prejuízo (R$)': pnl
        })

    df_out = pd.DataFrame(out_rows)
    
    print("\n===========================================================")
    print("     RELATÓRIO CONSOLIDADO DO ARSENAL COMPLETO - AGOSTO")
    print("===========================================================")
    print(f"Total de Operações: {len(df_out)}")
    tot_res = greens_count + reds_count
    wr_total = (greens_count / tot_res * 100) if tot_res > 0 else 0.0
    print(f"Resolvidas: {tot_res} | Greens: {greens_count} ({wr_total:.1f}%) | Reds: {reds_count} | Pendentes: {pending_count}")
    print(f"Lucro Líquido Acumulado (Stake R$ 100): R$ {pnl_total:,.2f}")
    print("-----------------------------------------------------------")
    for m, stats in by_method.items():
        g = stats['GREEN']
        rc = stats['RED']
        t_m = g + rc
        wr_m = (g / t_m * 100) if t_m > 0 else 0.0
        print(f"* {m}: {g}G / {rc}R ({wr_m:.1f}% Win Rate) -> P&L: R$ {stats['PNL']:,.2f}")
    print("===========================================================\n")

    # Formatação openpyxl
    file_paths = [
        'Sinais_Arsenal_Completo_Paper_Trading_01_a_17_Ago.xlsx',
        'lay0x3/Sinais_Arsenal_Completo_Paper_Trading_01_a_17_Ago.xlsx'
    ]

    for path in file_paths:
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Arsenal_Completo_Paper_Trading'

            headers = list(df_out.columns)
            ws.append(headers)

            header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            input_fill = PatternFill(start_color='FFF2CC', fill_type='solid')
            green_fill = PatternFill(start_color='D1FAE5', fill_type='solid')
            red_fill = PatternFill(start_color='FEE2E2', fill_type='solid')

            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )

            for r_idx, row_data in enumerate(df_out.values, 2):
                ws.append(list(row_data))
                for c_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center' if c_idx in [1, 2, 6, 7, 8, 9, 10, 11] else 'left', vertical='center')
                    
                    res_val = str(ws.cell(row=r_idx, column=10).value)
                    if c_idx in [8, 9, 10, 11]:
                        if res_val == 'GREEN':
                            cell.fill = green_fill
                        elif res_val == 'RED':
                            cell.fill = red_fill
                        else:
                            cell.fill = input_fill

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            wb.save(path)
            print(f"Planilha atualizada com sucesso em: {path}")
        except Exception as e:
            print(f"Erro ao salvar em {path}: {e}")

if __name__ == "__main__":
    atualizar_planilha_arsenal()
