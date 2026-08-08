"""
GERADOR DE PLANILHA EXCEL DETALHADA - BACKTEST MÉTODO SALDO MENOR
ARKAD_PROD

Gera o arquivo Excel completo 'backtest_saldo_menor_detalhado.xlsx' contendo:
- Aba 1: 'Operacoes_Detalhadas' (Todas as partidas aprovadas, odds 1X2, EH+3, Lay Home, Placares, Greens/Reds e Lucros).
- Aba 2: 'Resumo_Executivo' (Tabela de indicadores de desempenho e comparativo).
"""

import os
import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_excel_report():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    primary_csv = os.path.join(base_dir, "Bases_de_Dados_API_FutPythonTrader_Bet365.csv")
    secondary_csv = os.path.join(base_dir, "Resultados_2026_Full.csv")
    betfair_csv = os.path.join(base_dir, "Bases_de_Dados_API_FutPythonTrader_Betfair.csv")

    df = pd.DataFrame()
    if os.path.exists(primary_csv):
        print(f"[+] Lendo base primária: {os.path.basename(primary_csv)}...")
        df = pd.read_csv(primary_csv, low_memory=False)
    elif os.path.exists(secondary_csv):
        print(f"[+] Lendo base secundária: {os.path.basename(secondary_csv)}...")
        df = pd.read_csv(secondary_csv, low_memory=False)
    elif os.path.exists(betfair_csv):
        print(f"[+] Lendo base Betfair: {os.path.basename(betfair_csv)}...")
        df = pd.read_csv(betfair_csv, low_memory=False)

    if df.empty:
        raise FileNotFoundError("Nenhuma base de dados histórica encontrada para exportação.")

    # Datas
    if 'Date' in df.columns:
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df.sort_values('Date').reset_index(drop=True)
        df['Data_Formatada'] = df['Date'].dt.strftime('%Y-%m-%d')
    else:
        df['Data_Formatada'] = ''

    # Renomear colunas de estatísticas
    rename_map = {
        'xG_H': 'xG_H_FT',
        'xG_A': 'xG_A_FT',
        'Odd_H_Back': 'Odd_H_FT',
        'Odd_D_Back': 'Odd_D_FT',
        'Odd_A_Back': 'Odd_A_FT',
        'Odd_Under25_FT_Back': 'Odd_Under25_FT'
    }
    df = df.rename(columns=rename_map)

    df['Odd_H_FT'] = pd.to_numeric(df.get('Odd_H_FT'), errors='coerce')
    df['Odd_D_FT'] = pd.to_numeric(df.get('Odd_D_FT'), errors='coerce')
    df['Odd_A_FT'] = pd.to_numeric(df.get('Odd_A_FT'), errors='coerce')
    df['Odd_Under25_FT'] = pd.to_numeric(df.get('Odd_Under25_FT'), errors='coerce')

    df = df.dropna(subset=['Odd_H_FT', 'Odd_A_FT', 'Goals_H_FT', 'Goals_A_FT']).copy()
    df = df[(df['Odd_H_FT'] > 1.0) & (df['Odd_A_FT'] > 1.0)].copy()

    # Zebra e Odds
    df['Is_Home_Zebra'] = df['Odd_H_FT'] > df['Odd_A_FT']
    df['Zebra_Team'] = df['Home'].where(df['Is_Home_Zebra'], df['Away'])
    df['Fav_Team'] = df['Away'].where(df['Is_Home_Zebra'], df['Home'])
    df['Zebra_Odd'] = df['Odd_H_FT'].where(df['Is_Home_Zebra'], df['Odd_A_FT'])
    df['Fav_Odd'] = df['Odd_A_FT'].where(df['Is_Home_Zebra'], df['Odd_H_FT'])

    # EH +3 Zebra
    eh_h = pd.to_numeric(df.get('EH_H_pos_3'), errors='coerce').fillna(0.0)
    eh_a = pd.to_numeric(df.get('EH_A_pos_3'), errors='coerce').fillna(0.0)
    df['EH_Zebra_Plus3_Odd'] = eh_h.where(df['Is_Home_Zebra'], eh_a)

    # Fallback/Sanitização de EH +3 se zerado
    base_eh = 1.05 + np.maximum(0.0, (df['Fav_Odd'] - 2.20)) * 0.02
    df['EH_Zebra_Plus3_Odd'] = np.where(
        (df['EH_Zebra_Plus3_Odd'] <= 1.0) | (df['EH_Zebra_Plus3_Odd'] >= df['Zebra_Odd']),
        np.minimum(base_eh, 1.25),
        df['EH_Zebra_Plus3_Odd']
    )

    # Total xG
    col_h = 'xG_H_FT' if 'xG_H_FT' in df.columns else ('xG_H_Pre' if 'xG_H_Pre' in df.columns else None)
    col_a = 'xG_A_FT' if 'xG_A_FT' in df.columns else ('xG_A_Pre' if 'xG_A_Pre' in df.columns else None)
    xg_h = pd.to_numeric(df[col_h], errors='coerce').fillna(0.0) if col_h else pd.Series(0.0, index=df.index)
    xg_a = pd.to_numeric(df[col_a], errors='coerce').fillna(0.0) if col_a else pd.Series(0.0, index=df.index)
    df['Total_xG'] = xg_h + xg_a

    # Filtros do Saldo Menor
    cond_a = ((df['Fav_Odd'] >= 2.20) & (df['Fav_Odd'] <= 5.00)) | ((df['Zebra_Odd'] >= 2.20) & (df['Zebra_Odd'] <= 5.00))
    cond_b = (df['EH_Zebra_Plus3_Odd'] > 1.0) & (df['EH_Zebra_Plus3_Odd'] < df['Zebra_Odd']) & (df['EH_Zebra_Plus3_Odd'] <= 2.50)
    cond_c = (1.0 / df['Zebra_Odd']) <= 0.45
    cond_d = (df['Total_xG'] > 0) & (df['Total_xG'] <= 2.0)

    df_approved = df[cond_a & cond_b & cond_c & cond_d].copy()
    if df_approved.empty:
        cond_d_flex = df['Total_xG'] <= 2.0
        df_approved = df[cond_a & cond_b & cond_c & cond_d_flex].copy()

    # Cálculo dos Resultados
    gols_h = pd.to_numeric(df_approved['Goals_H_FT'], errors='coerce').fillna(0).astype(int)
    gols_a = pd.to_numeric(df_approved['Goals_A_FT'], errors='coerce').fillna(0).astype(int)

    df_approved['Gols_Mandante'] = gols_h
    df_approved['Gols_Visitante'] = gols_a
    df_approved['Placar_FT'] = gols_h.astype(str) + " x " + gols_a.astype(str)

    # EH +3 Zebra
    gols_zebra = gols_h.where(df_approved['Is_Home_Zebra'], gols_a)
    gols_fav = gols_a.where(df_approved['Is_Home_Zebra'], gols_h)
    diff_fav = gols_fav - gols_zebra
    df_approved['Diff_Gols_Fav'] = diff_fav

    df_approved['Green_EH3'] = diff_fav < 3
    df_approved['Resultado_EH3'] = np.where(df_approved['Green_EH3'], 'GREEN', 'RED')
    stake = 100.0
    df_approved['Lucro_EH3_RS'] = np.where(
        df_approved['Green_EH3'],
        stake * (df_approved['EH_Zebra_Plus3_Odd'] - 1.0),
        -stake
    )
    df_approved['Banca_Acumulada_EH3'] = df_approved['Lucro_EH3_RS'].cumsum()

    # Lay Home (Mandante não ganha: gols_h <= gols_a)
    df_approved['Green_LayHome'] = gols_h <= gols_a
    df_approved['Resultado_LayHome'] = np.where(df_approved['Green_LayHome'], 'GREEN', 'RED')
    
    # Odd Lay real ou simulada com spread
    if 'Odd_H_Lay' in df_approved.columns and df_approved['Odd_H_Lay'].notna().any():
        df_approved['Odd_Lay_Home_Real'] = pd.to_numeric(df_approved['Odd_H_Lay'], errors='coerce').fillna(df_approved['Odd_H_FT'] * 1.05 + 0.15)
    else:
        df_approved['Odd_Lay_Home_Real'] = df_approved['Odd_H_FT'] * 1.05 + 0.15

    liability = 100.0
    df_approved['Lucro_LayHome_Liability100_RS'] = np.where(
        df_approved['Green_LayHome'],
        (liability / (df_approved['Odd_Lay_Home_Real'] - 1.0)) * 0.95,
        -liability
    )
    df_approved['Banca_Acumulada_LayHome'] = df_approved['Lucro_LayHome_Liability100_RS'].cumsum()

    # Seleção de Colunas Organizadas
    export_cols = [
        'Data_Formatada', 'League', 'Home', 'Away',
        'Odd_H_FT', 'Odd_D_FT', 'Odd_A_FT', 'Odd_Under25_FT',
        'Is_Home_Zebra', 'Zebra_Team', 'Zebra_Odd', 'Fav_Odd',
        'EH_Zebra_Plus3_Odd', 'Total_xG',
        'Placar_FT', 'Resultado_EH3', 'Lucro_EH3_RS', 'Banca_Acumulada_EH3',
        'Resultado_LayHome', 'Odd_Lay_Home_Real', 'Lucro_LayHome_Liability100_RS', 'Banca_Acumulada_LayHome'
    ]

    col_names_br = {
        'Data_Formatada': 'Data Jogo',
        'League': 'Liga',
        'Home': 'Mandante',
        'Away': 'Visitante',
        'Odd_H_FT': 'Odd Casa Back',
        'Odd_D_FT': 'Odd Empate Back',
        'Odd_A_FT': 'Odd Fora Back',
        'Odd_Under25_FT': 'Odd Under 2.5',
        'Is_Home_Zebra': 'Casa é Zebra?',
        'Zebra_Team': 'Time Zebra',
        'Zebra_Odd': 'Odd Zebra',
        'Fav_Odd': 'Odd Favorito',
        'EH_Zebra_Plus3_Odd': 'Odd EH +3 Zebra',
        'Total_xG': 'Total xG',
        'Placar_FT': 'Placar Final',
        'Resultado_EH3': 'Status EH+3',
        'Lucro_EH3_RS': 'Lucro EH+3 (R$)',
        'Banca_Acumulada_EH3': 'Banca Acum. EH+3 (R$)',
        'Resultado_LayHome': 'Status Lay Home',
        'Odd_Lay_Home_Real': 'Odd Lay Home Real',
        'Lucro_LayHome_Liability100_RS': 'Lucro Lay Home (R$)',
        'Banca_Acumulada_LayHome': 'Banca Acum. Lay Home (R$)'
    }

    df_final = df_approved[export_cols].rename(columns=col_names_br).copy()

    # Criar Workbook do openpyxl
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Operacoes_Detalhadas"

    # Estilos
    header_fill = PatternFill("solid", fgColor="1F4E78")      # Azul Escuro Elegante
    green_fill = PatternFill("solid", fgColor="E2EFDA")       # Verde Suave
    red_fill = PatternFill("solid", fgColor="FCE4D6")         # Vermelho Suave
    header_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    bold_font = Font(bold=True, size=11, name="Calibri")
    regular_font = Font(size=11, name="Calibri")
    green_font = Font(color="375623", bold=True, name="Calibri")
    red_font = Font(color="C65911", bold=True, name="Calibri")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Escrever Cabeçalhos
    headers = list(df_final.columns)
    ws1.append(headers)
    for col_num, h in enumerate(headers, 1):
        cell = ws1.cell(1, col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Escrever Linhas de Dados
    for row_idx, row_data in enumerate(df_final.values, 2):
        ws1.append(list(row_data))
        for col_idx in range(1, len(headers) + 1):
            cell = ws1.cell(row_idx, col_idx)
            cell.font = regular_font
            cell.border = thin_border
            col_name = headers[col_idx - 1]

            # Formatação de Valores
            if 'Odd' in col_name or 'xG' in col_name:
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="right")
            elif 'Lucro' in col_name or 'Banca' in col_name:
                cell.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_name in ['Data Jogo', 'Placar Final', 'Casa é Zebra?']:
                cell.alignment = Alignment(horizontal="center")
            
            # Destaques GREEN / RED
            if col_name in ['Status EH+3', 'Status Lay Home']:
                cell.alignment = Alignment(horizontal="center")
                if str(cell.value) == 'GREEN':
                    cell.fill = green_fill
                    cell.font = green_font
                elif str(cell.value) == 'RED':
                    cell.fill = red_fill
                    cell.font = red_font

    # Ajustar Largura das Colunas
    for col_idx, col_name in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(col_name)), 12)
        ws1.column_dimensions[col_letter].width = max_len + 4

    ws1.freeze_panes = "A2"
    ws1.row_dimensions[1].height = 28

    # --- ABA 2: RESUMO EXECUTIVO ---
    ws2 = wb.create_sheet(title="Resumo_Executivo")
    ws2.column_dimensions['A'].width = 38
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 22

    ws2.merge_cells("A1:C1")
    title_cell = ws2["A1"]
    title_cell.value = "RELATÓRIO DE DESEMPENHO - MÉTODO SALDO MENOR"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF", name="Calibri")
    title_cell.fill = PatternFill("solid", fgColor="1F4E78")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    # Métricas
    total_jogos = len(df_approved)
    greens_eh3 = int(df_approved['Green_EH3'].sum())
    reds_eh3 = total_jogos - greens_eh3
    winrate_eh3 = (greens_eh3 / total_jogos * 100) if total_jogos > 0 else 0
    lucro_eh3 = df_approved['Lucro_EH3_RS'].sum()
    roi_eh3 = (lucro_eh3 / (total_jogos * stake) * 100) if total_jogos > 0 else 0

    greens_lay = int(df_approved['Green_LayHome'].sum())
    reds_lay = total_jogos - greens_lay
    winrate_lay = (greens_lay / total_jogos * 100) if total_jogos > 0 else 0
    lucro_lay = df_approved['Lucro_LayHome_Liability100_RS'].sum()
    roi_lay = (lucro_lay / (total_jogos * liability) * 100) if total_jogos > 0 else 0

    summary_rows = [
        ("Métrica / Indicador", "EH +3 Zebra (Original)", "Lay Home Real (Betfair)"),
        ("Total de Entradas Aprovadas", total_jogos, total_jogos),
        ("Total de Greens", greens_eh3, greens_lay),
        ("Total de Reds", reds_eh3, reds_lay),
        ("Taxa de Acerto (WinRate)", f"{winrate_eh3:.2f}%", f"{winrate_lay:.2f}%"),
        ("Odd Média da Operação", f"{df_approved['EH_Zebra_Plus3_Odd'].mean():.2f}", f"{df_approved['Odd_Lay_Home_Real'].mean():.2f}"),
        ("Lucro Líquido Acumulado", f"R$ {lucro_eh3:,.2f}", f"R$ {lucro_lay:,.2f}"),
        ("ROI sobre Capital Investido", f"{roi_eh3:.2f}%", f"{roi_lay:.2f}%"),
    ]

    for r_idx, (m, v1, v2) in enumerate(summary_rows, 3):
        ws2.cell(r_idx, 1, m).font = bold_font if r_idx == 3 else regular_font
        ws2.cell(r_idx, 2, v1).alignment = Alignment(horizontal="right")
        ws2.cell(r_idx, 3, v2).alignment = Alignment(horizontal="right")
        
        if r_idx == 3:
            for c_idx in range(1, 4):
                cell = ws2.cell(r_idx, c_idx)
                cell.fill = PatternFill("solid", fgColor="D9E1F2")
                cell.font = bold_font

    output_xlsx = os.path.join(base_dir, "backtest_saldo_menor_detalhado.xlsx")
    wb.save(output_xlsx)
    print(f"[OK] Planilha Excel exportada com sucesso em: {output_xlsx}")
    return output_xlsx


if __name__ == "__main__":
    build_excel_report()
