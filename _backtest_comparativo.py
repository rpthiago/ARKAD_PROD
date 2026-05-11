"""
BACKTEST COMPARATIVO — ANTES vs DEPOIS das melhorias
=====================================================
ANTES  = config atual (odd_max 11.5, 41 ligas 0x1, 3 ligas 1x0)
DEPOIS = melhorias aplicadas:
  - odd_max 0x1: 11.5 -> 13.0
  - Novas ligas 0x1: GERMANY 1, GERMANY 2, ITALY 2, ITALY 3, AUSTRIA 1,
                     SWITZERLAND 1, CZECH 1, FRANCE 3, EUROPA CHAMPIONS LEAGUE
  - Rodo ITALY 1 | Lay_CS_1x0 | Odd 10-12 REMOVIDO (2026 mostra 100% green)
  - SPAIN 1 rodos (odd <=6 e 6-8) MANTIDOS (dados historicos negativos)
  - Cap diario: 4 -> 6 apostas/metodo
"""
import json
from pathlib import Path
import pandas as pd
import numpy as np

ROOT    = Path("C:/Users/thiag/OneDrive/Documentos/GitHub/ARKAD_PROD")
BASE_CSV = Path("C:/Users/thiag/OneDrive/Documentos/GitHub/DASHBOARD_ARKAD-1/Bases_de_Dados_API_FutPythonTrader_Bet365.csv")
OUTPUT_DIR = ROOT / "Apostas_Diarias"
OUTPUT_DIR.mkdir(exist_ok=True)

BANCA_INICIAL    = 1100.0
RESPONSABILIDADE = 0.03
COMISSAO         = 0.065

# ── Carregar config ───────────────────────────────────────────────────────────
cfg = json.loads((ROOT / "config_universo_97.json").read_text(encoding="utf-8"))
fm  = cfg["runtime_data"]["filtros_metodo"]

# ── CENÁRIO ANTES ─────────────────────────────────────────────────────────────
ANTES = {
    "0x1": {
        "ligas":   {l.upper() for l in fm["Lay_CS_0x1_B365"]["ligas_permitidas"]},
        "odd_min": fm["Lay_CS_0x1_B365"]["odd_min"],
        "odd_max": fm["Lay_CS_0x1_B365"]["odd_max"],
        "cap":     4,
    },
    "1x0": {
        "ligas":   {l.upper() for l in fm["Lay_CS_1x0_B365"]["ligas_permitidas"]},
        "odd_min": fm["Lay_CS_1x0_B365"]["odd_min"],
        "odd_max": fm["Lay_CS_1x0_B365"]["odd_max"],
        "cap":     4,
    },
}

# ── CENÁRIO DEPOIS ────────────────────────────────────────────────────────────
NOVAS_LIGAS_0x1 = {
    "GERMANY 1", "GERMANY 2", "ITALY 2", "ITALY 3",
    "AUSTRIA 1", "SWITZERLAND 1", "CZECH 1",
    "FRANCE 3", "EUROPA CHAMPIONS LEAGUE",
}

DEPOIS = {
    "0x1": {
        "ligas":   ANTES["0x1"]["ligas"] | NOVAS_LIGAS_0x1,
        "odd_min": 8.0,
        "odd_max": 13.0,   # ampliado de 11.5 -> 13.0
        "cap":     6,      # cap diario aumentado
    },
    "1x0": {
        "ligas":   ANTES["1x0"]["ligas"],
        "odd_min": 4.5,
        "odd_max": 11.5,
        "cap":     6,
    },
}

# ── Rodos (mantidos do config) ────────────────────────────────────────────────
def _carregar_rodos(remover_ids=None):
    cuts, seen = [], set()
    for src in [
        cfg.get("filtros_rodo", []),
        cfg.get("filters", {}).get("filtros_rodo", []),
    ]:
        for c in (src or []):
            cid = c.get("id")
            if cid not in seen:
                seen.add(cid)
                if remover_ids and cid in remover_ids:
                    continue
                cuts.append(c)
    return cuts

RODOS_ANTES  = _carregar_rodos()
RODOS_DEPOIS = _carregar_rodos(remover_ids={11})  # remove rodo ITALY 1 | 1x0 | Odd 10-12

def _is_rodo(liga, metodo, odd, rodos):
    for cut in rodos:
        cut_leagues = set(cut.get("leagues", []))
        if cut.get("league"):
            cut_leagues.add(str(cut["league"]).upper())
        if cut_leagues and liga.upper() not in cut_leagues:
            continue
        me = cut.get("method_equals")
        mc = cut.get("method_contains")
        if me and str(me) != metodo:
            continue
        if mc and str(mc) not in metodo:
            continue
        omn = cut.get("odd_min")
        omx = cut.get("odd_max")
        if omn is not None and odd < float(omn):
            continue
        if omx is not None and odd > float(omx):
            continue
        return True
    return False

# ── Carregar base ─────────────────────────────────────────────────────────────
print("Carregando base Bet365...")
df = pd.read_csv(BASE_CSV, low_memory=False)
df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
df2026 = df[df["Date"].dt.year == 2026].dropna(subset=["Goals_H_FT", "Goals_A_FT"]).copy()
df2026["GH"]   = df2026["Goals_H_FT"].astype(int)
df2026["GA"]   = df2026["Goals_A_FT"].astype(int)
df2026["Liga"] = df2026["League"].str.upper().str.strip()
df2026["Jogo"] = df2026["Home"].str.strip() + " x " + df2026["Away"].str.strip()
df2026 = df2026.sort_values("Date").reset_index(drop=True)
print(f"OK: {len(df2026)} jogos 2026\n")

# ── Função de backtest ────────────────────────────────────────────────────────
def run_backtest(cenario, rodos, label):
    cfg_0x1 = cenario["0x1"]
    cfg_1x0 = cenario["1x0"]

    def filtrar(odd_col, odd_min, odd_max, ligas, metodo):
        s = df2026.copy()
        s["Odd"] = pd.to_numeric(s[odd_col], errors="coerce")
        s = s.dropna(subset=["Odd"])
        s = s[(s["Odd"] >= odd_min) & (s["Odd"] <= odd_max)].copy()
        if ligas:
            s = s[s["Liga"].isin(ligas)].copy()
        mask = s.apply(lambda r: _is_rodo(r["Liga"], metodo, r["Odd"], rodos), axis=1)
        s = s[~mask].copy()
        s["Metodo"]   = metodo
        s["Odd_Base"] = s["Odd"]
        return s

    df_0x1 = filtrar("Odd_CS_0x1", cfg_0x1["odd_min"], cfg_0x1["odd_max"], cfg_0x1["ligas"], "Lay_CS_0x1_B365")
    df_0x1["Green"] = ~((df_0x1["GH"] == 0) & (df_0x1["GA"] == 1))

    df_1x0 = filtrar("Odd_CS_1x0", cfg_1x0["odd_min"], cfg_1x0["odd_max"], cfg_1x0["ligas"], "Lay_CS_1x0_B365")
    df_1x0["Green"] = ~((df_1x0["GH"] == 1) & (df_1x0["GA"] == 0))

    df_all = pd.concat([df_0x1, df_1x0], ignore_index=True)
    df_all = df_all.sort_values(["Date", "Metodo"]).reset_index(drop=True)

    banca = BANCA_INICIAL
    registros = []

    for data, grupo in df_all.groupby(df_all["Date"].dt.date):
        g0 = grupo[grupo["Metodo"] == "Lay_CS_0x1_B365"].head(cfg_0x1["cap"])
        g1 = grupo[grupo["Metodo"] == "Lay_CS_1x0_B365"].head(cfg_1x0["cap"])
        dia = pd.concat([g0, g1])

        for _, row in dia.iterrows():
            odd   = float(row["Odd_Base"])
            green = bool(row["Green"])
            stake = banca * RESPONSABILIDADE
            lucro = (stake / (odd - 1)) * (1 - COMISSAO) if green else -stake
            banca += lucro
            registros.append({
                "Data":   str(data),
                "Liga":   str(row["Liga"]),
                "Metodo": str(row["Metodo"]),
                "Odd":    round(odd, 2),
                "Stake":  round(stake, 2),
                "Green":  "GREEN" if green else "RED",
                "Lucro":  round(lucro, 2),
                "Banca":  round(banca, 2),
            })

    df_ops = pd.DataFrame(registros)
    total  = len(df_ops)
    greens = (df_ops["Green"] == "GREEN").sum()
    lucro_total = banca - BANCA_INICIAL
    roi_total   = lucro_total / BANCA_INICIAL * 100
    banca_series = df_ops["Banca"].values
    pico  = np.maximum.accumulate(banca_series)
    dd_max = ((pico - banca_series) / pico * 100).max()

    # Por mes
    df_ops["Mes"] = pd.to_datetime(df_ops["Data"]).dt.to_period("M").astype(str)
    meses = []
    bm = BANCA_INICIAL
    for mes, grp in df_ops.groupby("Mes"):
        n = len(grp); g = (grp["Green"] == "GREEN").sum(); l = grp["Lucro"].sum()
        bm += l
        meses.append({"Mes": mes, "N": n, "Green%": round(g/n*100,1), "Lucro": round(l,2), "Banca": round(bm,2)})

    # Por metodo
    metodos = []
    for m, grp in df_ops.groupby("Metodo"):
        n = len(grp); g = (grp["Green"] == "GREEN").sum(); l = grp["Lucro"].sum()
        roi_m = l / grp["Stake"].sum() * 100 if grp["Stake"].sum() > 0 else 0
        metodos.append({"Metodo": m, "N": n, "Green%": round(g/n*100,1), "Lucro": round(l,2), "ROI%": round(roi_m,1)})

    return {
        "label":       label,
        "total":       total,
        "greens":      int(greens),
        "reds":        int(total - greens),
        "green_pct":   round(greens/total*100, 2) if total > 0 else 0,
        "banca_final": round(banca, 2),
        "lucro":       round(lucro_total, 2),
        "roi":         round(roi_total, 2),
        "dd_max":      round(dd_max, 1),
        "meses_pos":   sum(1 for m in meses if m["Lucro"] > 0),
        "total_meses": len(meses),
        "meses":       meses,
        "metodos":     metodos,
        "df_ops":      df_ops,
        "ligas_0x1":   len(cenario["0x1"]["ligas"]),
        "ligas_1x0":   len(cenario["1x0"]["ligas"]),
        "odd_max_0x1": cenario["0x1"]["odd_max"],
        "cap":         cenario["0x1"]["cap"],
    }

# ── Rodar os dois cenários ────────────────────────────────────────────────────
print("Rodando ANTES...")
r_antes  = run_backtest(ANTES,  RODOS_ANTES,  "ANTES  (config atual)")
print("Rodando DEPOIS...")
r_depois = run_backtest(DEPOIS, RODOS_DEPOIS, "DEPOIS (melhorias)")

# ── Relatório comparativo ─────────────────────────────────────────────────────
print()
print("=" * 75)
print("BACKTEST COMPARATIVO — ANTES vs DEPOIS")
print("=" * 75)

def linha(label, v_antes, v_depois, fmt="{}", melhor="maior"):
    diff = ""
    try:
        va = float(str(v_antes).replace("%","").replace("R$","").strip())
        vd = float(str(v_depois).replace("%","").replace("R$","").strip())
        delta = vd - va
        if melhor == "maior":
            sinal = "+" if delta > 0 else ""
            cor   = "MELHOR" if delta > 0 else ("PIOR" if delta < 0 else "=")
        else:
            sinal = "+" if delta > 0 else ""
            cor   = "PIOR" if delta > 0 else ("MELHOR" if delta < 0 else "=")
        diff = f"  [{sinal}{delta:.1f}] {cor}"
    except Exception:
        pass
    print(f"  {label:<30} {str(v_antes):<20} {str(v_depois):<20} {diff}")

print(f"\n  {'Indicador':<30} {'ANTES':<20} {'DEPOIS':<20} {'Diferenca'}")
print("  " + "-" * 70)
linha("Banca Final (R$)",    r_antes["banca_final"], r_depois["banca_final"])
linha("Lucro Total (R$)",    r_antes["lucro"],       r_depois["lucro"])
linha("ROI Total (%)",       r_antes["roi"],         r_depois["roi"])
linha("Total Apostas",       r_antes["total"],       r_depois["total"])
linha("Greens",              r_antes["greens"],      r_depois["greens"])
linha("Reds",                r_antes["reds"],        r_depois["reds"],        melhor="menor")
linha("Taxa Green (%)",      r_antes["green_pct"],   r_depois["green_pct"])
linha("Drawdown Max (%)",    r_antes["dd_max"],      r_depois["dd_max"],      melhor="menor")
linha("Meses Positivos",     f"{r_antes['meses_pos']}/{r_antes['total_meses']}", f"{r_depois['meses_pos']}/{r_depois['total_meses']}")
linha("Ligas 0x1",           r_antes["ligas_0x1"],   r_depois["ligas_0x1"])
linha("Odd Max 0x1",         r_antes["odd_max_0x1"], r_depois["odd_max_0x1"])
linha("Cap Diario",          r_antes["cap"],         r_depois["cap"])

print(f"\nRESULTADO POR MES — ANTES:")
for m in r_antes["meses"]:
    e = "OK" if m["Lucro"] > 0 else "XX"
    print(f"  [{e}] {m['Mes']}  N={m['N']:3d} | Green {m['Green%']}% | Lucro R$ {m['Lucro']:+.2f} | Banca R$ {m['Banca']:.2f}")

print(f"\nRESULTADO POR MES — DEPOIS:")
for m in r_depois["meses"]:
    e = "OK" if m["Lucro"] > 0 else "XX"
    print(f"  [{e}] {m['Mes']}  N={m['N']:3d} | Green {m['Green%']}% | Lucro R$ {m['Lucro']:+.2f} | Banca R$ {m['Banca']:.2f}")

print(f"\nRESULTADO POR METODO — ANTES:")
for m in r_antes["metodos"]:
    e = "OK" if m["ROI%"] > 0 else "XX"
    print(f"  [{e}] {m['Metodo']:<30} N={m['N']:3d} | Green {m['Green%']}% | ROI {m['ROI%']:+.1f}% | Lucro R$ {m['Lucro']:+.2f}")

print(f"\nRESULTADO POR METODO — DEPOIS:")
for m in r_depois["metodos"]:
    e = "OK" if m["ROI%"] > 0 else "XX"
    print(f"  [{e}] {m['Metodo']:<30} N={m['N']:3d} | Green {m['Green%']}% | ROI {m['ROI%']:+.1f}% | Lucro R$ {m['Lucro']:+.2f}")

# ── Exportar Excel comparativo ────────────────────────────────────────────────
output_path = OUTPUT_DIR / "Backtest_Comparativo_Melhorias_2026.xlsx"
print(f"\nExportando Excel: {output_path}")

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    # Aba comparativo
    comp = {
        "Indicador": [
            "Banca Final (R$)", "Lucro Total (R$)", "ROI Total (%)",
            "Total Apostas", "Greens", "Reds", "Taxa Green (%)",
            "Drawdown Max (%)", "Meses Positivos", "Ligas 0x1",
            "Odd Max 0x1", "Cap Diario",
        ],
        "ANTES": [
            r_antes["banca_final"], r_antes["lucro"], r_antes["roi"],
            r_antes["total"], r_antes["greens"], r_antes["reds"], r_antes["green_pct"],
            r_antes["dd_max"], f"{r_antes['meses_pos']}/{r_antes['total_meses']}",
            r_antes["ligas_0x1"], r_antes["odd_max_0x1"], r_antes["cap"],
        ],
        "DEPOIS": [
            r_depois["banca_final"], r_depois["lucro"], r_depois["roi"],
            r_depois["total"], r_depois["greens"], r_depois["reds"], r_depois["green_pct"],
            r_depois["dd_max"], f"{r_depois['meses_pos']}/{r_depois['total_meses']}",
            r_depois["ligas_0x1"], r_depois["odd_max_0x1"], r_depois["cap"],
        ],
    }
    df_comp = pd.DataFrame(comp)
    df_comp["Diferenca"] = df_comp.apply(
        lambda row: f"{float(str(row['DEPOIS']).split('/')[0]) - float(str(row['ANTES']).split('/')[0]):+.1f}"
        if str(row["ANTES"]).replace(".","").replace("-","").replace("+","").split("/")[0].isdigit()
        else "", axis=1
    )
    df_comp.to_excel(writer, sheet_name="Comparativo", index=False)

    # Operacoes ANTES
    r_antes["df_ops"].to_excel(writer, sheet_name="Operacoes_ANTES", index=False)

    # Operacoes DEPOIS
    r_depois["df_ops"].to_excel(writer, sheet_name="Operacoes_DEPOIS", index=False)

    # Por mes
    df_mes = pd.DataFrame([
        {"Mes": m["Mes"], "Apostas_ANTES": m["N"], "Green_ANTES": m["Green%"], "Lucro_ANTES": m["Lucro"], "Banca_ANTES": m["Banca"]}
        for m in r_antes["meses"]
    ])
    df_mes2 = pd.DataFrame([
        {"Mes": m["Mes"], "Apostas_DEPOIS": m["N"], "Green_DEPOIS": m["Green%"], "Lucro_DEPOIS": m["Lucro"], "Banca_DEPOIS": m["Banca"]}
        for m in r_depois["meses"]
    ])
    df_mes_full = df_mes.merge(df_mes2, on="Mes", how="outer")
    df_mes_full.to_excel(writer, sheet_name="Por_Mes", index=False)

    # Novas ligas adicionadas
    novas = sorted(NOVAS_LIGAS_0x1)
    pd.DataFrame({"Nova_Liga_0x1": novas}).to_excel(writer, sheet_name="Novas_Ligas", index=False)

    # Formatacao
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    for sname in writer.sheets:
        ws = writer.sheets[sname]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F3460")
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

    # Colorir comparativo
    ws_c = writer.sheets["Comparativo"]
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill   = PatternFill("solid", fgColor="FFC7CE")
    headers = [c.value for c in ws_c[1]]
    if "Diferenca" in headers:
        col_d = headers.index("Diferenca") + 1
        for row_n in range(2, ws_c.max_row + 1):
            val = str(ws_c.cell(row=row_n, column=col_d).value or "")
            if val.startswith("+"):
                ws_c.cell(row=row_n, column=col_d).fill = green_fill
            elif val.startswith("-"):
                ws_c.cell(row=row_n, column=col_d).fill = red_fill

print(f"Excel salvo: {output_path}")
print("\nConcluido!")
