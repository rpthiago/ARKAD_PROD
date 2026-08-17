"""
GERADOR DE BACKTEST COMPLETO DO MÉTODO LAY 2X2 - MÊS DE AGOSTO DE 2026 (01 a 17/08)
ARKAD_PROD

Este script compila todas as partidas do mês de Agosto (01/08 a 17/08/2026) que atenderam
aos critérios quantitativos da estratégia Lay 2x2, cruzando dados estáticos e os endpoints diários da API.
"""

import os
import shutil
import difflib
from pathlib import Path
import pandas as pd
import numpy as np
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from futpythontrader_client import get_daily_dataframe
from metodo_lay2x2_strategy import validar_entrada_lay2x2, calcular_resultado_lay2x2

print("==========================================================================")
print("     PROCESSANDO BACKTEST COMPLETO DO LAY 2X2 - AGOSTO (01 A 17/08)")
print("==========================================================================\n")

STAKE_UNIDADE = 100.0
COMISSAO_BETFAIR = 0.05

def norm_str(s: str) -> str:
    return ''.join(c for c in str(s).lower() if c.isalnum())

# 1. Construir base de placares da ESPN API e DBs locais para 01 a 17/08
db_scores = {}

# ESPN API
for day in range(1, 18):
    date_str = f"202608{day:02d}"
    url = f"https://site.api.espn.com/apis/site/v2/sports/soccer/all/scoreboard?dates={date_str}&limit=1000"
    try:
        r = requests.get(url, timeout=10)
        if r.status_code == 200:
            data = r.json()
            events = data.get('events', [])
            for ev in events:
                try:
                    competitions = ev.get('competitions', [])[0]
                    competitors = competitions.get('competitors', [])
                    home_comp = [c for c in competitors if c.get('homeAway') == 'home'][0]
                    away_comp = [c for c in competitors if c.get('homeAway') == 'away'][0]
                    
                    h_name = home_comp.get('team', {}).get('displayName', '')
                    a_name = away_comp.get('team', {}).get('displayName', '')
                    
                    status_type = competitions.get('status', {}).get('type', {}).get('name', '')
                    if status_type in ['STATUS_FULL_TIME', 'STATUS_FINAL', 'STATUS_APPROVED', 'FULL_TIME']:
                        h_score = int(home_comp.get('score', 0))
                        a_score = int(away_comp.get('score', 0))
                        
                        dt_fmt = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
                        hk = norm_str(h_name)
                        ak = norm_str(a_name)
                        db_scores[(dt_fmt, hk, ak)] = (h_score, a_score)
                        db_scores[(dt_fmt, hk[:5], ak[:5])] = (h_score, a_score)
                except Exception:
                    pass
    except Exception:
        pass

# DBs estáticos locais
for fpath in ['Bases_de_Dados_API_FutPythonTrader_Betfair_FRESH.csv', 'Bases_de_Dados_API_FutPythonTrader_Bet365.csv']:
    if Path(fpath).exists():
        try:
            df_s = pd.read_csv(fpath, low_memory=False)
            d_col = [c for c in df_s.columns if 'date' in c.lower() or 'data' in c.lower()][0]
            h_col = 'Home' if 'Home' in df_s.columns else ('Home_Team' if 'Home_Team' in df_s.columns else 'Mandante')
            a_col = 'Away' if 'Away' in df_s.columns else ('Away_Team' if 'Away_Team' in df_s.columns else 'Visitante')
            gh_col = [c for c in df_s.columns if 'goals_h_ft' in c.lower() or 'gols mandante' in c.lower()][0]
            ga_col = [c for c in df_s.columns if 'goals_a_ft' in c.lower() or 'gols visitante' in c.lower()][0]
            
            df_s['d_str'] = pd.to_datetime(df_s[d_col], errors='coerce').dt.strftime('%Y-%m-%d')
            for _, r in df_s.iterrows():
                dt = r['d_str']
                if pd.notna(dt) and dt >= '2026-08-01':
                    gh_v = r[gh_col]
                    ga_v = r[ga_col]
                    if pd.notna(gh_v) and pd.notna(ga_v):
                        try:
                            hk = norm_str(r[h_col])
                            ak = norm_str(r[a_col])
                            db_scores[(dt, hk, ak)] = (int(float(gh_v)), int(float(ga_v)))
                            db_scores[(dt, hk[:5], ak[:5])] = (int(float(gh_v)), int(float(ga_v)))
                        except Exception:
                            pass
        except Exception:
            pass

print(f"Base integrada de placares com {len(db_scores)} partidas finalizadas.")

# 2. Carregar todos os jogos de Agosto (01 a 07 static + 08 a 17 daily API)
df_bf = pd.read_csv("Bases_de_Dados_API_FutPythonTrader_Betfair_FRESH.csv", low_memory=False)
df_bf["d_str"] = pd.to_datetime(df_bf["Date"], errors="coerce").dt.strftime("%Y-%m-%d")
aug_static = df_bf[(df_bf["d_str"] >= "2026-08-01") & (df_bf["d_str"] <= "2026-08-07")].copy()

all_dfs = [aug_static]

for day in range(8, 18):
    d_str = f"2026-08-{day:02d}"
    try:
        df_d = get_daily_dataframe("betfair", d_str)
        if not df_d.empty:
            df_d["d_str"] = d_str
            all_dfs.append(df_d)
    except Exception:
        pass

df_full_aug = pd.concat(all_dfs, ignore_index=True)
print(f"Total de partidas obtidas em Agosto (01-17/08): {len(df_full_aug)}")

df_full_aug["Odd_CS_2x2_Lay"] = pd.to_numeric(df_full_aug["Odd_CS_2x2_Lay"], errors="coerce")
df_full_aug["Odd_Under25"] = pd.to_numeric(df_full_aug.get("Odd_Under25_FT_Back", df_full_aug.get("Odd_Under25")), errors="coerce")
df_full_aug["Odd_H"] = pd.to_numeric(df_full_aug.get("Odd_H_FT_Back", df_full_aug.get("Odd_H")), errors="coerce")
df_full_aug["Odd_A"] = pd.to_numeric(df_full_aug.get("Odd_A_FT_Back", df_full_aug.get("Odd_A")), errors="coerce")

rows_out = []
greens_count = 0
reds_count = 0
pending_count = 0
pnl_stake100_total = 0.0
pnl_liab200_total = 0.0

for idx, r in df_full_aug.iterrows():
    odd_lay_2x2 = r["Odd_CS_2x2_Lay"]
    odd_u25 = r["Odd_Under25"]
    odd_h = r["Odd_H"]
    odd_a = r["Odd_A"]
    
    if pd.notna(odd_lay_2x2) and 8.0 <= odd_lay_2x2 <= 18.0:
        dt = str(r["d_str"])
        time_str = str(r.get("Time", r.get("horario", "15:00")))[:5] if pd.notna(r.get("Time", r.get("horario"))) else "15:00"
        liga = str(r.get("League", r.get("Div", r.get("liga", "Desconhecida"))))
        home = str(r.get("Home", r.get("Home_Team", r.get("Mandante", ""))))
        away = str(r.get("Away", r.get("Away_Team", r.get("Visitante", ""))))
        confronto = f"{home} x {away}"
        
        gh, ga = None, None
        
        # Tenta pegar placar original da linha
        if 'Goals_H_FT' in r and pd.notna(r['Goals_H_FT']) and 'Goals_A_FT' in r and pd.notna(r['Goals_A_FT']):
            try:
                gh = int(float(r['Goals_H_FT']))
                ga = int(float(r['Goals_A_FT']))
            except Exception:
                pass
                
        # Se não tiver na linha, busca no db_scores
        if gh is None or ga is None:
            hk = norm_str(home)
            ak = norm_str(away)
            if (dt, hk, ak) in db_scores:
                gh, ga = db_scores[(dt, hk, ak)]
            elif (dt, hk[:5], ak[:5]) in db_scores:
                gh, ga = db_scores[(dt, hk[:5], ak[:5])]
            else:
                for (dt_k, hk_k, ak_k), sc in db_scores.items():
                    if dt_k == dt and len(hk_k) > 4:
                        sim_h = difflib.SequenceMatcher(None, hk, hk_k).ratio()
                        sim_a = difflib.SequenceMatcher(None, ak, ak_k).ratio()
                        if sim_h >= 0.45 and sim_a >= 0.45:
                            gh, ga = sc
                            break
                            
        res_str = ""
        pnl_stk = ""
        pnl_liab = ""
        
        if gh is not None and ga is not None:
            is_2x2 = (gh == 2 and ga == 2)
            win = not is_2x2
            res_str = "GREEN" if win else "RED"
            
            pnl_stk = round(STAKE_UNIDADE * (1.0 - COMISSAO_BETFAIR), 2) if win else -round((odd_lay_2x2 - 1.0) * STAKE_UNIDADE, 2)
            stake_liab = 200.0 / (odd_lay_2x2 - 1.0)
            pnl_liab = round(stake_liab * (1.0 - COMISSAO_BETFAIR), 2) if win else -200.0
            
            if win:
                greens_count += 1
            else:
                reds_count += 1
                
            pnl_stake100_total += pnl_stk
            pnl_liab200_total += pnl_liab
        else:
            pending_count += 1
            
        rows_out.append({
            "Data": dt,
            "Horário": time_str,
            "Liga": liga,
            "Confronto": confronto,
            "Método": "Lay 2x2 Quant",
            "Lado": "LAY",
            "Odd Lay 2x2 Betfair": odd_lay_2x2,
            "Gols Mandante": gh if gh is not None else "",
            "Gols Visitante": ga if ga is not None else "",
            "Resultado (GREEN/RED)": res_str,
            "Lucro/Prejuízo (Stake R$100)": pnl_stk,
            "Lucro/Prejuízo (Liab R$200)": pnl_liab,
            "Justificativa": f"Odd Lay 2x2 ({odd_lay_2x2:.2f}) aprovada na faixa de risco 8.0 - 18.0"
        })

df_out = pd.DataFrame(rows_out)

print(f"\n===========================================================")
print(f"     BACKTEST COMPLETO DO LAY 2X2 (01 A 17 AGOSTO)")
print(f"===========================================================")
print(f"Total de Oportunidades Aprovadas no Mês: {len(df_out)}")
tot_res = greens_count + reds_count
wr_total = (greens_count / tot_res * 100) if tot_res > 0 else 0.0
print(f"Partidas Finalizadas Resolvidas: {tot_res}")
print(f"Greens: {greens_count} ({wr_total:.1f}%) | Reds: {reds_count} | Pendentes: {pending_count}")
print(f"Lucro Acumulado (Stake Fixa R$ 100): R$ {pnl_stake100_total:,.2f}")
print(f"Lucro Acumulado (Responsabilidade Fixa R$ 200): R$ {pnl_liab200_total:,.2f}")
print("===========================================================\n")

# 3. Gerar Excel formatado
out_file_root = "Backtest_Lay_2x2_Agosto_2026.xlsx"
out_file_lay = "lay0x3/Backtest_Lay_2x2_Agosto_2026.xlsx"

for path in [out_file_root, out_file_lay]:
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Backtest_Lay_2x2_Agosto"

        headers = list(df_out.columns)
        ws.append(headers)

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        input_fill = PatternFill(start_color="FFF2CC", fill_type="solid")
        green_fill = PatternFill(start_color="D1FAE5", fill_type="solid")
        red_fill = PatternFill(start_color="FEE2E2", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        for r_idx, row_data in enumerate(df_out.values, 2):
            ws.append(list(row_data))
            for c_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if c_idx in [1, 2, 6, 7, 8, 9, 10, 11, 12] else "left", vertical="center")
                
                res_val = str(ws.cell(row=r_idx, column=10).value)
                if c_idx in [8, 9, 10, 11, 12]:
                    if res_val == "GREEN":
                        cell.fill = green_fill
                    elif res_val == "RED":
                        cell.fill = red_fill
                    else:
                        cell.fill = input_fill

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(path)
        print(f"Planilha salva com sucesso em: {path}")
    except Exception as e:
        print(f"Aviso ao salvar em {path}: {e}")

print("\nProcessamento do Backtest Completo de Agosto para Lay 2x2 concluído!")
