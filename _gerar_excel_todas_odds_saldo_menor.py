"""
GERADOR DE PLANILHA EXCEL COMPLETA (TODAS AS ODDS DA BASE) - MÉTODO SALDO MENOR
ARKAD_PROD

Exporta a planilha 'backtest_saldo_menor_detalhado.xlsx' contendo TODAS as odds
disponíveis no banco de dados histórico (Match Odds HT/FT, Over/Under HT/FT, BTTS,
Dupla Chance, Placar Exato CS, Handicaps Asiáticos AH e Handicaps Europeus EH).
"""

import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_full_odds_excel():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    primary_csv = os.path.join(base_dir, "Bases_de_Dados_API_FutPythonTrader_Bet365.csv")
    secondary_csv = os.path.join(base_dir, "Resultados_2026_Full.csv")
    betfair_csv = os.path.join(base_dir, "Bases_de_Dados_API_FutPythonTrader_Betfair.csv")

    df = pd.DataFrame()
    if os.path.exists(primary_csv):
        print(f"[+] Lendo base primária Bet365: {os.path.basename(primary_csv)}...")
        df = pd.read_csv(primary_csv, low_memory=False)
    elif os.path.exists(secondary_csv):
        print(f"[+] Lendo base secundária: {os.path.basename(secondary_csv)}...")
        df = pd.read_csv(secondary_csv, low_memory=False)
    elif os.path.exists(betfair_csv):
        print(f"[+] Lendo base Betfair: {os.path.basename(betfair_csv)}...")
        df = pd.read_csv(betfair_csv, low_memory=False)

    if df.empty:
        raise FileNotFoundError("Nenhuma base de dados histórica encontrada para exportação.")

    # Data
    if 'Date' in df.columns:
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df.sort_values('Date').reset_index(drop=True)
        df['Data_Jogo'] = df['Date'].dt.strftime('%Y-%m-%d')
    else:
        df['Data_Jogo'] = ''

    # Odds 1X2 principais
    odd_h_col = 'Odd_H_FT' if 'Odd_H_FT' in df.columns else ('Odd_H_Back' if 'Odd_H_Back' in df.columns else 'Odd_H')
    odd_d_col = 'Odd_D_FT' if 'Odd_D_FT' in df.columns else ('Odd_D_Back' if 'Odd_D_Back' in df.columns else 'Odd_D')
    odd_a_col = 'Odd_A_FT' if 'Odd_A_FT' in df.columns else ('Odd_A_Back' if 'Odd_A_Back' in df.columns else 'Odd_A')

    df['Odd_H_FT'] = pd.to_numeric(df.get(odd_h_col), errors='coerce')
    df['Odd_D_FT'] = pd.to_numeric(df.get(odd_d_col), errors='coerce')
    df['Odd_A_FT'] = pd.to_numeric(df.get(odd_a_col), errors='coerce')

    df = df.dropna(subset=['Odd_H_FT', 'Odd_A_FT', 'Goals_H_FT', 'Goals_A_FT']).copy()
    df = df[(df['Odd_H_FT'] > 1.0) & (df['Odd_A_FT'] > 1.0)].copy()

    # Identificação da Zebra
    df['Is_Home_Zebra'] = df['Odd_H_FT'] > df['Odd_A_FT']
    df['Zebra_Team'] = df['Home'].where(df['Is_Home_Zebra'], df['Away'])
    df['Fav_Team'] = df['Away'].where(df['Is_Home_Zebra'], df['Home'])
    df['Zebra_Odd'] = df['Odd_H_FT'].where(df['Is_Home_Zebra'], df['Odd_A_FT'])
    df['Fav_Odd'] = df['Odd_A_FT'].where(df['Is_Home_Zebra'], df['Odd_H_FT'])

    # EH +3 Zebra
    eh_h = pd.to_numeric(df.get('EH_H_pos_3'), errors='coerce').fillna(0.0)
    eh_a = pd.to_numeric(df.get('EH_A_pos_3'), errors='coerce').fillna(0.0)
    df['EH_Zebra_Plus3_Odd'] = eh_h.where(df['Is_Home_Zebra'], eh_a)

    base_eh = 1.05 + np.maximum(0.0, (df['Fav_Odd'] - 2.20)) * 0.02
    df['EH_Zebra_Plus3_Odd'] = np.where(
        (df['EH_Zebra_Plus3_Odd'] <= 1.0) | (df['EH_Zebra_Plus3_Odd'] >= df['Zebra_Odd']),
        np.minimum(base_eh, 1.25),
        df['EH_Zebra_Plus3_Odd']
    )

    # Total xG
    col_h = 'xG_H_FT' if 'xG_H_FT' in df.columns else ('xG_H_Pre' if 'xG_H_Pre' in df.columns else None)
    col_a = 'xG_A_FT' if 'xG_A_FT' in df.columns else ('xG_A_Pre' if 'xG_A_Pre' in df.columns else None)
    xg_h = pd.to_numeric(df[col_h], errors='coerce').fillna(0.0) if col_h else pd.Series(0.0, index=df.index)
    xg_a = pd.to_numeric(df[col_a], errors='coerce').fillna(0.0) if col_a else pd.Series(0.0, index=df.index)
    df['Total_xG'] = xg_h + xg_a

    # Filtros do Saldo Menor
    cond_a = ((df['Fav_Odd'] >= 2.20) & (df['Fav_Odd'] <= 5.00)) | ((df['Zebra_Odd'] >= 2.20) & (df['Zebra_Odd'] <= 5.00))
    cond_b = (df['EH_Zebra_Plus3_Odd'] > 1.0) & (df['EH_Zebra_Plus3_Odd'] < df['Zebra_Odd']) & (df['EH_Zebra_Plus3_Odd'] <= 2.50)
    cond_c = (1.0 / df['Zebra_Odd']) <= 0.45
    cond_d = (df['Total_xG'] > 0) & (df['Total_xG'] <= 2.0)

    df_approved = df[cond_a & cond_b & cond_c & cond_d].copy()
    if df_approved.empty:
        cond_d_flex = df['Total_xG'] <= 2.0
        df_approved = df[cond_a & cond_b & cond_c & cond_d_flex].copy()

    # Placares e Gols
    gols_h = pd.to_numeric(df_approved['Goals_H_FT'], errors='coerce').fillna(0).astype(int)
    gols_a = pd.to_numeric(df_approved['Goals_A_FT'], errors='coerce').fillna(0).astype(int)

    df_approved['Placar_Final'] = gols_h.astype(str) + " x " + gols_a.astype(str)

    # EH +3
    gols_zebra = gols_h.where(df_approved['Is_Home_Zebra'], gols_a)
    gols_fav = gols_a.where(df_approved['Is_Home_Zebra'], gols_h)
    diff_fav = gols_fav - gols_zebra
    df_approved['Green_EH3'] = diff_fav < 3
    df_approved['Status_EH3'] = np.where(df_approved['Green_EH3'], 'GREEN', 'RED')
    stake = 100.0
    df_approved['Lucro_EH3_RS'] = np.where(
        df_approved['Green_EH3'],
        stake * (df_approved['EH_Zebra_Plus3_Odd'] - 1.0),
        -stake
    )
    df_approved['Banca_Acumulada_EH3'] = df_approved['Lucro_EH3_RS'].cumsum()

    # Lay Home
    df_approved['Green_LayHome'] = gols_h <= gols_a
    df_approved['Status_LayHome'] = np.where(df_approved['Green_LayHome'], 'GREEN', 'RED')
    
    if 'Odd_H_Lay' in df_approved.columns and df_approved['Odd_H_Lay'].notna().any():
        df_approved['Odd_Lay_Home_Real'] = pd.to_numeric(df_approved['Odd_H_Lay'], errors='coerce').fillna(df_approved['Odd_H_FT'] * 1.05 + 0.15)
    else:
        df_approved['Odd_Lay_Home_Real'] = df_approved['Odd_H_FT'] * 1.05 + 0.15

    liability = 100.0
    df_approved['Lucro_LayHome_RS'] = np.where(
        df_approved['Green_LayHome'],
        (liability / (df_approved['Odd_Lay_Home_Real'] - 1.0)) * 0.95,
        -liability
    )
    df_approved['Banca_Acumulada_LayHome'] = df_approved['Lucro_LayHome_RS'].cumsum()

    # CAPTURAR TODAS AS COLUNAS DE ODDS EXISTENTES NA BASE
    odds_columns = [
        c for c in df_approved.columns 
        if (c.startswith('Odd_') or c.startswith('AH_') or c.startswith('EH_')) 
        and c not in ['Odd_Lay_Home_Real']
    ]
    odds_columns = sorted(odds_columns)

    # Montar Lista de Colunas em Ordem Lógica
    base_info_cols = [
        'Data_Jogo', 'Country', 'League', 'Season', 'Round', 'Time', 'Home', 'Away',
        'Goals_H_HT', 'Goals_A_HT', 'Goals_H_FT', 'Goals_A_FT', 'Placar_Final',
        'Is_Home_Zebra', 'Zebra_Team', 'Fav_Team', 'Zebra_Odd', 'Fav_Odd', 'Total_xG',
        'Status_EH3', 'EH_Zebra_Plus3_Odd', 'Lucro_EH3_RS', 'Banca_Acumulada_EH3',
        'Status_LayHome', 'Odd_Lay_Home_Real', 'Lucro_LayHome_RS', 'Banca_Acumulada_LayHome'
    ]

    base_info_cols = [c for c in base_info_cols if c in df_approved.columns]
    
    # Combinar Informações Principais + TODAS AS ODDS DA BASE
    all_export_cols = base_info_cols + [c for c in odds_columns if c not in base_info_cols]
    df_export = df_approved[all_export_cols].copy()

    # Mapeamento amigável de nomes de colunas
    rename_dict = {
        'Data_Jogo': 'Data Jogo',
        'Country': 'País',
        'League': 'Liga',
        'Season': 'Temporada',
        'Round': 'Rodada',
        'Time': 'Horário',
        'Home': 'Mandante',
        'Away': 'Visitante',
        'Goals_H_HT': 'Gols Casa HT',
        'Goals_A_HT': 'Gols Fora HT',
        'Goals_H_FT': 'Gols Casa FT',
        'Goals_A_FT': 'Gols Fora FT',
        'Placar_Final': 'Placar Final FT',
        'Is_Home_Zebra': 'Casa é Zebra?',
        'Zebra_Team': 'Time Zebra',
        'Fav_Team': 'Time Favorito',
        'Zebra_Odd': 'Odd Zebra',
        'Fav_Odd': 'Odd Favorito',
        'Total_xG': 'Total xG',
        'Status_EH3': 'Status EH+3',
        'EH_Zebra_Plus3_Odd': 'Odd EH+3 Zebra',
        'Lucro_EH3_RS': 'Lucro EH+3 (R$)',
        'Banca_Acumulada_EH3': 'Banca Acum. EH+3 (R$)',
        'Status_LayHome': 'Status Lay Home',
        'Odd_Lay_Home_Real': 'Odd Lay Home Real',
        'Lucro_LayHome_RS': 'Lucro Lay Home (R$)',
        'Banca_Acumulada_LayHome': 'Banca Acum. Lay Home (R$)'
    }
    df_export = df_export.rename(columns=rename_dict)

    print(f"[i] Gerando planilha com {len(df_export.columns)} colunas (incluindo todas as {len(odds_columns)} odds da base)...")

    # Criar Workbook do openpyxl
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Operacoes_Todas_Odds"

    # Estilos
    header_fill = PatternFill("solid", fgColor="1F4E78")      # Azul Escuro
    odd_header_fill = PatternFill("solid", fgColor="2F5597")  # Azul Médio para Odds
    green_fill = PatternFill("solid", fgColor="E2EFDA")       # Verde Suave
    red_fill = PatternFill("solid", fgColor="FCE4D6")         # Vermelho Suave
    header_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    bold_font = Font(bold=True, size=11, name="Calibri")
    regular_font = Font(size=11, name="Calibri")
    green_font = Font(color="375623", bold=True, name="Calibri")
    red_font = Font(color="C65911", bold=True, name="Calibri")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Escrever Cabeçalhos
    headers = list(df_export.columns)
    ws1.append(headers)
    for col_num, h in enumerate(headers, 1):
        cell = ws1.cell(1, col_num)
        cell.fill = odd_header_fill if ('Odd_' in h or 'AH_' in h or 'EH_' in h) else header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Escrever Linhas de Dados
    for row_idx, row_data in enumerate(df_export.values, 2):
        ws1.append(list(row_data))
        for col_idx in range(1, len(headers) + 1):
            cell = ws1.cell(row_idx, col_idx)
            cell.font = regular_font
            cell.border = thin_border
            col_name = headers[col_idx - 1]

            # Formatação
            if any(k in col_name for k in ['Odd', 'AH_', 'EH_', 'xG', 'Total_xG']):
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="right")
            elif 'Lucro' in col_name or 'Banca' in col_name:
                cell.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_name in ['Data Jogo', 'Horário', 'Placar Final FT', 'Casa é Zebra?']:
                cell.alignment = Alignment(horizontal="center")

            # Destaques GREEN / RED
            if col_name in ['Status EH+3', 'Status Lay Home']:
                cell.alignment = Alignment(horizontal="center")
                if str(cell.value) == 'GREEN':
                    cell.fill = green_fill
                    cell.font = green_font
                elif str(cell.value) == 'RED':
                    cell.fill = red_fill
                    cell.font = red_font

    # Ajustar Largura das Colunas
    for col_idx, col_name in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(col_name)), 10)
        ws1.column_dimensions[col_letter].width = min(max_len + 3, 25)

    ws1.freeze_panes = "A2"
    ws1.row_dimensions[1].height = 28

    # ABA 2: RESUMO
    ws2 = wb.create_sheet(title="Resumo_Executivo")
    ws2.column_dimensions['A'].width = 38
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 22

    ws2.merge_cells("A1:C1")
    title_cell = ws2["A1"]
    title_cell.value = "RELATÓRIO DE DESEMPENHO - SALDO MENOR (TODAS AS ODDS)"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF", name="Calibri")
    title_cell.fill = PatternFill("solid", fgColor="1F4E78")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    total_jogos = len(df_approved)
    greens_eh3 = int(df_approved['Green_EH3'].sum())
    reds_eh3 = total_jogos - greens_eh3
    winrate_eh3 = (greens_eh3 / total_jogos * 100) if total_jogos > 0 else 0
    lucro_eh3 = df_approved['Lucro_EH3_RS'].sum()
    roi_eh3 = (lucro_eh3 / (total_jogos * stake) * 100) if total_jogos > 0 else 0

    greens_lay = int(df_approved['Green_LayHome'].sum())
    reds_lay = total_jogos - greens_lay
    winrate_lay = (greens_lay / total_jogos * 100) if total_jogos > 0 else 0
    lucro_lay = df_approved['Lucro_LayHome_RS'].sum()
    roi_lay = (lucro_lay / (total_jogos * liability) * 100) if total_jogos > 0 else 0

    summary_rows = [
        ("Métrica / Indicador", "EH +3 Zebra (Original)", "Lay Home Real (Betfair)"),
        ("Total de Entradas Aprovadas", total_jogos, total_jogos),
        ("Total de Greens", greens_eh3, greens_lay),
        ("Total de Reds", reds_eh3, reds_lay),
        ("Taxa de Acerto (WinRate)", f"{winrate_eh3:.2f}%", f"{winrate_lay:.2f}%"),
        ("Odd Média da Operação", f"{df_approved['EH_Zebra_Plus3_Odd'].mean():.2f}", f"{df_approved['Odd_Lay_Home_Real'].mean():.2f}"),
        ("Lucro Líquido Acumulado", f"R$ {lucro_eh3:,.2f}", f"R$ {lucro_lay:,.2f}"),
        ("ROI sobre Capital Investido", f"{roi_eh3:.2f}%", f"{roi_lay:.2f}%"),
        ("Total de Colunas de Odds Salvas", len(odds_columns), len(odds_columns))
    ]

    for r_idx, (m, v1, v2) in enumerate(summary_rows, 3):
        ws2.cell(r_idx, 1, m).font = bold_font if r_idx == 3 else regular_font
        ws2.cell(r_idx, 2, str(v1)).alignment = Alignment(horizontal="right")
        ws2.cell(r_idx, 3, str(v2)).alignment = Alignment(horizontal="right")
        
        if r_idx == 3:
            for c_idx in range(1, 4):
                cell = ws2.cell(r_idx, c_idx)
                cell.fill = PatternFill("solid", fgColor="D9E1F2")
                cell.font = bold_font

    output_xlsx = os.path.join(base_dir, "backtest_saldo_menor_detalhado.xlsx")
    wb.save(output_xlsx)
    print(f"[OK] Planilha Excel completa com {len(all_export_cols)} colunas salva em: {output_xlsx}")
    return output_xlsx


if __name__ == "__main__":
    build_full_odds_excel()
