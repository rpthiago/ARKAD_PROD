"""
BACKTEST ARKAD PROD ELITE 2026
================================
Metodos: Lay_CS_0x1_B365 e Lay_CS_1x0_B365
Filtros: ligas_permitidas + rodos toxicos (blacklist) do config_universo_97.json
Gestao: responsabilidade fixa 3% da banca | cap diario 4 apostas/metodo
Periodo: 2026-01-01 -> hoje

NOTA sobre regra dupla:
  - 0x1 opera em 41 ligas (sem SPAIN/ITALY)
  - 1x0 opera em 3 ligas (SPAIN 1, ITALY 1, SPAIN 2)
  - Nao ha intersecao de ligas, entao cada metodo opera INDEPENDENTE
  - A regra dupla do main.py so se aplica quando o mesmo jogo aparece
    nas duas listas (raro, pois ligas diferentes)
"""

import json
from pathlib import Path

import pandas as pd
import numpy as np

# ── Caminhos ──────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent
CFG_PATH = ROOT / "config_universo_97.json"
BASE_CSV = Path("C:/Users/thiag/OneDrive/Documentos/GitHub/DASHBOARD_ARKAD-1/Bases_de_Dados_API_FutPythonTrader_Bet365.csv")
OUTPUT_DIR = ROOT / "Apostas_Diarias"
OUTPUT_DIR.mkdir(exist_ok=True)

# ── Parametros de gestao ──────────────────────────────────────────────────────
BANCA_INICIAL    = 1100.0
RESPONSABILIDADE = 0.03    # 3% da banca
CAP_DIARIO       = 4       # max apostas por metodo por dia
COMISSAO         = 0.065   # 6.5% Betfair

# ── Carregar config ───────────────────────────────────────────────────────────
cfg = json.loads(CFG_PATH.read_text(encoding="utf-8"))
FILTROS_METODO = cfg.get("runtime_data", {}).get("filtros_metodo", {})

LIGAS_0x1   = {l.upper() for l in FILTROS_METODO.get("Lay_CS_0x1_B365", {}).get("ligas_permitidas", [])}
LIGAS_1x0   = {l.upper() for l in FILTROS_METODO.get("Lay_CS_1x0_B365", {}).get("ligas_permitidas", [])}
ODD_MIN_0x1 = FILTROS_METODO.get("Lay_CS_0x1_B365", {}).get("odd_min", 8.0)
ODD_MAX_0x1 = FILTROS_METODO.get("Lay_CS_0x1_B365", {}).get("odd_max", 11.5)
ODD_MIN_1x0 = FILTROS_METODO.get("Lay_CS_1x0_B365", {}).get("odd_min", 4.5)
ODD_MAX_1x0 = FILTROS_METODO.get("Lay_CS_1x0_B365", {}).get("odd_max", 11.5)

# ── Rodos toxicos (blacklist) ─────────────────────────────────────────────────
def _carregar_rodos():
    cuts, seen = [], set()
    for src in [
        cfg.get("filtros_rodo", []),
        cfg.get("filters", {}).get("filtros_rodo", []),
        cfg.get("filters", {}).get("toxic_cuts", []),
    ]:
        for c in (src or []):
            cid = c.get("id")
            if cid not in seen:
                seen.add(cid)
                cuts.append(c)
    return cuts

RODOS = _carregar_rodos()

def _is_rodo(liga: str, metodo: str, odd: float) -> bool:
    for cut in RODOS:
        cut_leagues = set(cut.get("leagues", []))
        if cut.get("league"):
            cut_leagues.add(str(cut["league"]).upper())
        if cut_leagues and liga.upper() not in cut_leagues:
            continue
        me = cut.get("method_equals")
        mc = cut.get("method_contains")
        if me and str(me) != metodo:
            continue
        if mc and str(mc) not in metodo:
            continue
        omn = cut.get("odd_min")
        omx = cut.get("odd_max")
        if omn is not None and odd < float(omn):
            continue
        if omx is not None and odd > float(omx):
            continue
        return True
    return False

# ── Carregar base ─────────────────────────────────────────────────────────────
print("=" * 70)
print("BACKTEST ARKAD PROD ELITE 2026")
print("=" * 70)
print(f"   Banca inicial: R$ {BANCA_INICIAL:.2f}")
print(f"   Responsabilidade: {RESPONSABILIDADE*100:.0f}% (stake dinamica)")
print(f"   Metodos: Lay_CS_0x1_B365 + Lay_CS_1x0_B365")
print(f"   Rodos toxicos: {len(RODOS)} combinacoes bloqueadas")
print(f"   Ligas 0x1: {len(LIGAS_0x1)} | Ligas 1x0: {len(LIGAS_1x0)}")
print()

print("Carregando base Bet365...")
df = pd.read_csv(BASE_CSV, low_memory=False)
df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
df2026 = df[df["Date"].dt.year == 2026].dropna(subset=["Goals_H_FT", "Goals_A_FT"]).copy()
df2026["GH"]   = df2026["Goals_H_FT"].astype(int)
df2026["GA"]   = df2026["Goals_A_FT"].astype(int)
df2026["Liga"] = df2026["League"].astype(str).str.upper().str.strip()
df2026["Jogo"] = df2026["Home"].astype(str) + " x " + df2026["Away"].astype(str)
df2026 = df2026.sort_values("Date").reset_index(drop=True)

print(f"OK: {len(df2026)} jogos com resultado em 2026 ({df2026['Date'].min().date()} -> {df2026['Date'].max().date()})")
print()

# ── Filtrar elegiveis por metodo ──────────────────────────────────────────────
def filtrar_metodo(df_in, metodo, odd_col, odd_min, odd_max, ligas_perm):
    sub = df_in.copy()
    sub[odd_col] = pd.to_numeric(sub[odd_col], errors="coerce")
    sub = sub.dropna(subset=[odd_col])
    sub = sub[(sub[odd_col] >= odd_min) & (sub[odd_col] <= odd_max)].copy()
    if ligas_perm:
        sub = sub[sub["Liga"].isin(ligas_perm)].copy()
    mask_rodo = sub.apply(lambda r: _is_rodo(r["Liga"], metodo, r[odd_col]), axis=1)
    sub = sub[~mask_rodo].copy()
    sub["Metodo"]   = metodo
    sub["Odd_Base"] = sub[odd_col]
    return sub

# Lay_CS_0x1_B365 — GREEN se placar final NAO for 0x1
df_0x1 = filtrar_metodo(df2026, "Lay_CS_0x1_B365", "Odd_CS_0x1", ODD_MIN_0x1, ODD_MAX_0x1, LIGAS_0x1)
df_0x1 = df_0x1.copy()
df_0x1["Green"] = ~((df_0x1["GH"] == 0) & (df_0x1["GA"] == 1))

# Lay_CS_1x0_B365 — GREEN se placar final NAO for 1x0
df_1x0 = filtrar_metodo(df2026, "Lay_CS_1x0_B365", "Odd_CS_1x0", ODD_MIN_1x0, ODD_MAX_1x0, LIGAS_1x0)
df_1x0 = df_1x0.copy()
df_1x0["Green"] = ~((df_1x0["GH"] == 1) & (df_1x0["GA"] == 0))

print(f"Elegiveis apos filtros (ligas + rodos):")
print(f"   Lay_CS_0x1_B365: {len(df_0x1)} apostas | Green: {df_0x1['Green'].mean()*100:.1f}%")
print(f"   Lay_CS_1x0_B365: {len(df_1x0)} apostas | Green: {df_1x0['Green'].mean()*100:.1f}%")
print()

# ── Combinar e ordenar ────────────────────────────────────────────────────────
df_all = pd.concat([df_0x1, df_1x0], ignore_index=True)
df_all = df_all.sort_values(["Date", "Metodo"]).reset_index(drop=True)

# ── Backtest com cap diario ───────────────────────────────────────────────────
print(f"Iniciando backtest — Banca: R$ {BANCA_INICIAL:.2f} | Responsabilidade: {RESPONSABILIDADE*100:.0f}%")
print("=" * 70)

banca = BANCA_INICIAL
registros = []

for data, grupo in df_all.groupby(df_all["Date"].dt.date):
    grupo_0x1 = grupo[grupo["Metodo"] == "Lay_CS_0x1_B365"].head(CAP_DIARIO)
    grupo_1x0 = grupo[grupo["Metodo"] == "Lay_CS_1x0_B365"].head(CAP_DIARIO)
    dia = pd.concat([grupo_0x1, grupo_1x0])

    for _, row in dia.iterrows():
        odd   = float(row["Odd_Base"])
        green = bool(row["Green"])
        stake = banca * RESPONSABILIDADE

        if green:
            lucro = (stake / (odd - 1)) * (1 - COMISSAO)
        else:
            lucro = -stake

        banca += lucro

        registros.append({
            "Data":   str(data),
            "Liga":   str(row["Liga"]),
            "Jogo":   str(row["Jogo"]),
            "Metodo": str(row["Metodo"]),
            "Odd":    round(odd, 2),
            "Stake":  round(stake, 2),
            "Green":  "GREEN" if green else "RED",
            "Lucro":  round(lucro, 2),
            "Banca":  round(banca, 2),
        })

df_ops = pd.DataFrame(registros)

# ── Estatisticas ──────────────────────────────────────────────────────────────
total      = len(df_ops)
greens     = (df_ops["Green"] == "GREEN").sum()
reds       = (df_ops["Green"] == "RED").sum()
taxa_green = greens / total * 100 if total > 0 else 0
lucro_total = banca - BANCA_INICIAL
roi_total   = lucro_total / BANCA_INICIAL * 100

banca_series = df_ops["Banca"].values
pico  = np.maximum.accumulate(banca_series)
dd    = (pico - banca_series) / pico * 100
dd_max = dd.max()

seq_green = seq_red = cur_g = cur_r = 0
for g in df_ops["Green"]:
    if g == "GREEN":
        cur_g += 1; cur_r = 0; seq_green = max(seq_green, cur_g)
    else:
        cur_r += 1; cur_g = 0; seq_red = max(seq_red, cur_r)

# Por metodo
stats_metodo = []
for m, grp in df_ops.groupby("Metodo"):
    n = len(grp)
    g = (grp["Green"] == "GREEN").sum()
    l = grp["Lucro"].sum()
    roi_m = l / grp["Stake"].sum() * 100 if grp["Stake"].sum() > 0 else 0
    stats_metodo.append({
        "Metodo": m, "Apostas": n, "Green": int(g),
        "Green%": round(g/n*100, 1), "Lucro": round(l, 2), "ROI%": round(roi_m, 1)
    })

# Por mes
df_ops["Mes"] = pd.to_datetime(df_ops["Data"]).dt.to_period("M").astype(str)
stats_mes, banca_mes = [], BANCA_INICIAL
for mes, grp in df_ops.groupby("Mes"):
    n = len(grp); g = (grp["Green"] == "GREEN").sum(); l = grp["Lucro"].sum()
    banca_mes += l
    stats_mes.append({"Mes": mes, "Apostas": n, "Green%": round(g/n*100,1), "Lucro": round(l,2), "Banca": round(banca_mes,2)})

meses_pos = sum(1 for s in stats_mes if s["Lucro"] > 0)

# Por liga
stats_liga = []
for liga, grp in df_ops.groupby("Liga"):
    n = len(grp); g = (grp["Green"] == "GREEN").sum(); l = grp["Lucro"].sum()
    roi_l = l / grp["Stake"].sum() * 100 if grp["Stake"].sum() > 0 else 0
    stats_liga.append({"Liga": liga, "Apostas": n, "Green%": round(g/n*100,1), "Lucro": round(l,2), "ROI%": round(roi_l,1)})

df_liga = pd.DataFrame(stats_liga).sort_values("ROI%", ascending=False)

# ── Relatorio ─────────────────────────────────────────────────────────────────
print()
print("=" * 70)
print("BACKTEST ARKAD PROD ELITE 2026 — RELATORIO COMPLETO")
print("=" * 70)

print(f"\nRESUMO GERAL:")
print(f"   {'Banca Inicial (R$)':<35} {BANCA_INICIAL}")
print(f"   {'Banca Final (R$)':<35} {banca:.2f}")
print(f"   {'Lucro Total (R$)':<35} {lucro_total:.2f}")
print(f"   {'ROI Total (%)':<35} {roi_total:.2f}")
print(f"   {'Total Apostas':<35} {total}")
print(f"   {'Greens':<35} {greens}")
print(f"   {'Reds':<35} {reds}")
print(f"   {'Taxa Green (%)':<35} {taxa_green:.2f}")
print(f"   {'Drawdown Maximo (%)':<35} {dd_max:.1f}")
print(f"   {'Seq. Max. Greens':<35} {seq_green}")
print(f"   {'Seq. Max. Reds':<35} {seq_red}")
print(f"   {'Banca Pico (R$)':<35} {max(banca_series):.2f}")
print(f"   {'Banca Vale (R$)':<35} {min(banca_series):.2f}")
print(f"   {'Meses Positivos':<35} {meses_pos}/{len(stats_mes)}")
print(f"   {'Periodo':<35} {df_ops['Data'].min()} -> {df_ops['Data'].max()}")

print(f"\nRESULTADO POR METODO:")
for s in sorted(stats_metodo, key=lambda x: -x["ROI%"]):
    emoji = "OK" if s["ROI%"] > 0 else "XX"
    print(f"   [{emoji}] {s['Metodo']:<30} ROI: {s['ROI%']:+.1f}% | Apostas: {s['Apostas']} | Green: {s['Green%']}% | Lucro: R$ {s['Lucro']:+.2f}")

print(f"\nRESULTADO POR MES:")
for s in stats_mes:
    emoji = "OK" if s["Lucro"] > 0 else "XX"
    print(f"   [{emoji}] {s['Mes']}  Apostas: {s['Apostas']} | Green: {s['Green%']}% | Lucro: R$ {s['Lucro']:+.2f} | Banca: R$ {s['Banca']:.2f}")

print(f"\nTOP 10 MELHORES LIGAS:")
for _, row in df_liga.head(10).iterrows():
    emoji = "OK" if row["ROI%"] > 0 else "XX"
    print(f"   [{emoji}] {row['Liga']:<40} ROI: {row['ROI%']:+.1f}% | Apostas: {row['Apostas']} | Lucro: R$ {row['Lucro']:+.2f}")

print(f"\nTOP 5 PIORES LIGAS:")
for _, row in df_liga.tail(5).iterrows():
    emoji = "XX" if row["ROI%"] < 0 else "OK"
    print(f"   [{emoji}] {row['Liga']:<40} ROI: {row['ROI%']:+.1f}% | Apostas: {row['Apostas']} | Lucro: R$ {row['Lucro']:+.2f}")

# ── Exportar Excel ────────────────────────────────────────────────────────────
output_path = OUTPUT_DIR / "Backtest_ArkadProd_Elite_2026.xlsx"
print(f"\nExportando Excel: {output_path}")

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df_ops.to_excel(writer, sheet_name="Operacoes", index=False)
    pd.DataFrame(stats_metodo).to_excel(writer, sheet_name="Por_Metodo", index=False)
    pd.DataFrame(stats_mes).to_excel(writer, sheet_name="Por_Mes", index=False)
    df_liga.to_excel(writer, sheet_name="Por_Liga", index=False)

    resumo_data = {
        "Indicador": [
            "Banca Inicial (R$)", "Banca Final (R$)", "Lucro Total (R$)", "ROI Total (%)",
            "Total Apostas", "Greens", "Reds", "Taxa Green (%)",
            "Drawdown Maximo (%)", "Seq. Max. Greens", "Seq. Max. Reds",
            "Banca Pico (R$)", "Banca Vale (R$)", "Meses Positivos", "Periodo",
        ],
        "Valor": [
            BANCA_INICIAL, round(banca, 2), round(lucro_total, 2), round(roi_total, 2),
            total, int(greens), int(reds), round(taxa_green, 2),
            round(dd_max, 1), seq_green, seq_red,
            round(float(max(banca_series)), 2), round(float(min(banca_series)), 2),
            f"{meses_pos}/{len(stats_mes)}",
            f"{df_ops['Data'].min()} -> {df_ops['Data'].max()}",
        ],
    }
    pd.DataFrame(resumo_data).to_excel(writer, sheet_name="Resumo", index=False)

    # Rodos bloqueados
    rodos_cols = ["id", "name", "league", "method_equals", "odd_min", "odd_max"]
    rodos_cols_exist = [c for c in rodos_cols if any(c in r for r in RODOS)]
    if RODOS:
        df_rodos = pd.DataFrame(RODOS)
        cols_ok = [c for c in rodos_cols if c in df_rodos.columns]
        df_rodos[cols_ok].to_excel(writer, sheet_name="Rodos_Bloqueados", index=False)

    # Formatacao
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    for sname in writer.sheets:
        ws = writer.sheets[sname]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F3460")
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

    # Colorir GREEN/RED
    ws_ops = writer.sheets["Operacoes"]
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill   = PatternFill("solid", fgColor="FFC7CE")
    headers = [c.value for c in ws_ops[1]]
    if "Green" in headers:
        col_g = headers.index("Green") + 1
        for row_n in range(2, ws_ops.max_row + 1):
            val  = ws_ops.cell(row=row_n, column=col_g).value
            fill = green_fill if val == "GREEN" else red_fill
            for col_n in range(1, ws_ops.max_column + 1):
                ws_ops.cell(row=row_n, column=col_n).fill = fill

print(f"Excel salvo: {output_path}")
print(f"\nConcluido!")
