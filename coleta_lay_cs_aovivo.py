"""
coleta_lay_cs_aovivo.py — coleta ao vivo de odds in-play do Lay 0x1 e Lay 1x0
==============================================================================
Igual ao coleta_layhome, mas para placar exato (Correct Score). A cada dia
adiciona os jogos sinalizados (odd do placar 0-1 / 1-0 entre 7 e 12, ligas
principais) em DUAS planilhas .xlsx (uma por mercado). Você anota, DO SEU SITE,
a odd do placar exato na ABERTURA (min 0), min 60 e min 75. Placar e momento
dos gols se preenchem sozinhos.

Uso:  python coleta_lay_cs_aovivo.py [YYYY-MM-DD]
"""
import os, sys, datetime, unicodedata, re, ast, time
import pandas as pd, numpy as np
sys.stdout.reconfigure(encoding="utf-8")

MAJORS = ["ENGLAND 1","SPAIN 1","ITALY 1","GERMANY 1","FRANCE 1","BRAZIL 1",
          "PORTUGAL 1","NETHERLANDS 1","ENGLAND 2","BRAZIL 2","ARGENTINA 1"]
COLS = ["Date","Horario","Liga","Mandante","Visitante","Metodo","Prob","Odd_lay_entrada",
        "PREENCHER_odd_abertura","PREENCHER_odd_min60","PREENCHER_odd_min75",
        "Placar_final","Momento_gols","status","obs"]
# Filtro do Top 5 (a pagina ignora o Decision estrito da estrategia e usa este):
PROB_MIN_TOP5, ODD_MIN_TOP5, ODD_MAX_TOP5 = 0.50, 6.0, 99.0
FILLC = {"PREENCHER_odd_abertura","PREENCHER_odd_min60","PREENCHER_odd_min75","obs"}
# mercado -> estrategias reais (modelo) que selecionam os jogos + odd lay + arquivo + placar
# Agora a coleta roda os MODELOS do Top 5 (nao um filtro de odd) para gravar exatamente
# os jogos que os metodos apostariam, marcando qual metodo pegou cada um.
MERCADOS = {
    "0x0": dict(strategies=[("lay_0x0_rf_v2_strategy","RF")],
                odd_key="Odd_CS_0x0_Lay", ledger="coleta_lay0x0_aovivo.xlsx", placar="0-0"),
    "0x1": dict(strategies=[("lay_0x1_agressivo_strategy","Trader"),("lay_0x1_rf_strategy","RF")],
                odd_key="Odd_CS_0x1_Lay", ledger="coleta_lay0x1_aovivo.xlsx", placar="0-1"),
    "1x0": dict(strategies=[("lay_1x0_agressivo_strategy","Trader")],
                odd_key="Odd_CS_1x0_Lay", ledger="coleta_lay1x0_aovivo.xlsx", placar="1-0"),
}
# O "Lay 0x1 RF" treina RandomForest on-the-fly (mais lento) — OK, a rotina roda 1x/dia.

def _canon(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii","ignore").decode().lower()
    return re.sub(r"[^a-z0-9]", "", s)

def carregar(ledger):
    if os.path.exists(ledger) and os.path.getsize(ledger) > 0:
        try: return pd.read_excel(ledger, sheet_name="Jogos", dtype=str)
        except Exception: pass
    return pd.DataFrame(columns=COLS)

def salvar_xlsx(df, ledger):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    widths = {"Date":11,"Horario":8,"Liga":13,"Mandante":18,"Visitante":18,"Metodo":13,"Prob":7,"Odd_lay_entrada":14,
              "PREENCHER_odd_abertura":16,"PREENCHER_odd_min60":14,"PREENCHER_odd_min75":14,
              "Placar_final":11,"Momento_gols":26,"status":10,"obs":14}
    for tent in range(6):
        try:
            wb=Workbook(); ws=wb.active; ws.title="Jogos"
            fh=PatternFill("solid",fgColor="305496"); fy=PatternFill("solid",fgColor="FFF2CC"); fr=PatternFill("solid",fgColor="E2EFDA")
            white=Font(color="FFFFFF",bold=True); thin=Border(*[Side(style="thin",color="D0D0D0")]*4)
            for j,c in enumerate(COLS,1):
                cell=ws.cell(1,j,c); cell.font=white; cell.fill=fh
                cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=thin
            for i,(_,row) in enumerate(df.iterrows(),2):
                for j,c in enumerate(COLS,1):
                    v=row.get(c,""); cell=ws.cell(i,j,"" if pd.isna(v) else v); cell.border=thin
                    if c in FILLC: cell.fill=fy
                    elif c=="Odd_lay_entrada": cell.fill=fr
            for j,c in enumerate(COLS,1): ws.column_dimensions[get_column_letter(j)].width=widths.get(c,12)
            ws.freeze_panes="A2"; ws.row_dimensions[1].height=28
            wb.save(ledger); return True
        except PermissionError:
            if tent==5: return False
            time.sleep(2)
    return False

_HISTDF = {}
def _hist_df():
    """Historico recente (ultimos ~450 dias) — a forma rolante so precisa dos jogos
    recentes de cada time, e trimar acelera MUITO o feature-building das estrategias.
    OK para coleta ao vivo (hoje/amanha); nao usar para backtest de data antiga."""
    if "df" not in _HISTDF:
        from b365_data_utils import load_b365_historical
        df = load_b365_historical().copy()
        dt = pd.to_datetime(df["Date"], errors="coerce")
        corte = dt.max() - pd.Timedelta(days=450)
        _HISTDF["df"] = df[dt >= corte].reset_index(drop=True)
    return _HISTDF["df"]

def sinais_do_dia(date_str, cfg, diag=None):
    """Roda os MODELOS reais do metodo (Trader/RF) sobre os jogos do dia e retorna
    exatamente os jogos que eles apostariam, marcando qual metodo pegou cada um.
    Se `diag` (dict) for passado, preenche: n_api (jogos que a API trouxe) e
    errors (falhas por estrategia) — para a UI distinguir 'sem jogos' de 'sem sinal'."""
    from b365_data_utils import fetch_betfair_daily
    bf = fetch_betfair_daily(date_str)
    if diag is not None:
        diag["n_api"] = 0 if (bf is None or bf.empty) else len(bf)
    if bf is None or bf.empty: return []
    payload = bf.to_dict("records")
    hist = _hist_df()
    picks = {}   # (Home,Away) -> {dados + set de metodos}
    for mod_name, tag in cfg["strategies"]:
        try:
            mod = __import__(mod_name, fromlist=["predict_and_evaluate_live"])
            res = mod.predict_and_evaluate_live(payload, hist)
        except Exception as e:
            if diag is not None:
                diag.setdefault("errors", []).append(f"[{tag}] {mod_name}: {e}")
            print(f"    [{tag}] {mod_name}: ERRO {str(e)[:80]}"); continue
        for g in (res or []):
            if cfg["placar"] == "0-0" and g.get("Decision") != "APOSTA":
                continue
            # MESMO filtro da pagina Top 5 (ignora o Decision estrito da estrategia):
            odd = pd.to_numeric(g.get(cfg["odd_key"]) or np.nan, errors="coerce")
            prob = pd.to_numeric(g.get("Prob_ML") or np.nan, errors="coerce")
            if pd.isna(odd) or odd < ODD_MIN_TOP5 or odd > ODD_MAX_TOP5: continue
            if pd.isna(prob) or prob < PROB_MIN_TOP5: continue
            home, away = str(g.get("Home","")), str(g.get("Away",""))
            key = (home, away)
            if key not in picks:
                picks[key] = dict(Home=home, Away=away, Liga=g.get("League","") or g.get("Liga",""),
                                  Horario=str(g.get("Time","") or ""), odd=odd, prob=prob, metodos=set())
            picks[key]["metodos"].add(tag)
            if pd.isna(picks[key]["odd"]) and pd.notna(odd): picks[key]["odd"]=odd
            if pd.notna(prob) and prob > picks[key]["prob"]: picks[key]["prob"]=prob
    out = []
    for p in picks.values():
        out.append(dict(Date=date_str, Horario=p["Horario"], Liga=p["Liga"],
                        Mandante=p["Home"], Visitante=p["Away"], Metodo="+".join(sorted(p["metodos"])),
                        Prob=round(float(p["prob"])*100,1) if pd.notna(p["prob"]) else "",
                        Odd_lay_entrada=round(float(p["odd"]),2) if pd.notna(p["odd"]) else "",
                        PREENCHER_odd_abertura="", PREENCHER_odd_min60="", PREENCHER_odd_min75="",
                        Placar_final="", Momento_gols="", status="AGUARDA", obs=""))
    return out

_HISTCACHE = {}
def _hist():
    if "h" not in _HISTCACHE:
        from b365_data_utils import load_b365_historical
        h = load_b365_historical().copy()
        h["d"] = pd.to_datetime(h["Date"], errors="coerce").dt.strftime("%Y-%m-%d")
        h["ch"] = h["Home"].map(_canon); h["ca"] = h["Away"].map(_canon)
        h["gh"] = pd.to_numeric(h.get("Goals_H_FT"), errors="coerce")
        h["ga"] = pd.to_numeric(h.get("Goals_A_FT"), errors="coerce")
        idx = {}
        for _, r in h.dropna(subset=["gh","ga"]).iterrows():
            idx[(r["d"], r["ch"], r["ca"])] = r
        _HISTCACHE["h"] = idx
    return _HISTCACHE["h"]

def _parse(v):
    try:
        s=str(v).strip(); return [int(str(x).split("+")[0]) for x in ast.literal_eval(s)] if s not in("","[]","nan") else []
    except: return []

def liquidar(led):
    pend = led[led["status"] == "AGUARDA"]
    if pend.empty: return led, 0
    idx = _hist(); n = 0
    for i in pend.index:
        key = (str(led.at[i,"Date"])[:10], _canon(led.at[i,"Mandante"]), _canon(led.at[i,"Visitante"]))
        r = idx.get(key)
        if r is None: continue
        gh, ga = int(r["gh"]), int(r["ga"])
        mh, ma = _parse(r.get("Goals_Min_H")), _parse(r.get("Goals_Min_A"))
        mom = " | ".join([p for p in ["Casa:"+",".join(map(str,sorted(mh))) if mh else "",
                                       "Fora:"+",".join(map(str,sorted(ma))) if ma else ""] if p]) or "(sem gols)"
        led.at[i,"Placar_final"] = f"{gh}-{ga}"; led.at[i,"Momento_gols"] = mom; led.at[i,"status"] = "ENCERRADO"; n += 1
    return led, n

def processa(nome, cfg, datas):
    print("\n" + "#"*64 + f"\n### LAY {nome}  (placar {cfg['placar']})\n" + "#"*64)
    led = carregar(cfg["ledger"])
    led, nliq = liquidar(led)
    if nliq: print(f"  ✓ {nliq} jogo(s) encerrado(s).")
    ja = set(zip(led["Date"].astype(str), led["Mandante"].astype(str), led["Visitante"].astype(str))) if not led.empty else set()
    tot = 0
    for date_str in datas:
        novos = [s for s in sinais_do_dia(date_str, cfg) if (s["Date"], s["Mandante"], s["Visitante"]) not in ja]
        if novos:
            led = pd.concat([led, pd.DataFrame(novos)], ignore_index=True) if not led.empty else pd.DataFrame(novos)
            ja |= {(s["Date"], s["Mandante"], s["Visitante"]) for s in novos}; tot += len(novos)
            print(f"  ✓ {len(novos)} jogo(s) em {date_str}.")
    if not tot: print(f"  nenhum jogo novo em {datas} (fora de temporada / sem sinais na faixa).")
    led = led.reindex(columns=COLS)
    if not salvar_xlsx(led, cfg["ledger"]):
        print("  ⚠️ planilha aberta no Excel — nao gravou. Feche e rode de novo.")
    falta = led[led["PREENCHER_odd_abertura"].fillna("")==""].sort_values("Date")
    print(f"  --- JOGOS LAY {nome} PARA ANOTAR (placar {cfg['placar']}: abertura + min60 + min75) ---")
    if falta.empty:
        print("     (nenhum pendente)")
    else:
        for _, r in falta.iterrows():
            tag = "a jogar" if r["status"]=="AGUARDA" else f"FIM {r['Placar_final']}"
            met = "" if pd.isna(r.get("Metodo")) else str(r.get("Metodo",""))
            odde = "" if pd.isna(r.get("Odd_lay_entrada")) else r.get("Odd_lay_entrada")
            print(f"     {r['Date']} {str(r.get('Horario',''))[:5]:>5}  [{met:<10}] {str(r['Mandante'])[:15]:<15} x {str(r['Visitante'])[:15]:<15} (odd {cfg['placar']} ~{odde}, {tag})")
    print(f"     arquivo: {cfg['ledger']}")

def main():
    import pytz
    args = [a for a in sys.argv[1:] if re.match(r"\d{4}-\d{2}-\d{2}", a)]
    hoje = datetime.datetime.now(pytz.timezone("America/Sao_Paulo"))
    datas = args if args else [hoje.strftime("%Y-%m-%d"), (hoje + datetime.timedelta(days=1)).strftime("%Y-%m-%d")]
    for nome, cfg in MERCADOS.items():
        processa(nome, cfg, datas)

if __name__ == "__main__":
    main()
