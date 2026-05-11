"""
Auditoria SUPER ELITE — ARKAD PROD
====================================
Página de auditoria de sinais com filtro SUPER ELITE:
  - Sinal só é ELITE se ROI Histórico > 10% E status = aprovado
  - Trava de segurança: EV+ calculado (odd_real > 1/prob_green_historica)
  - Exibe no máximo os TOP 10 sinais do dia, ordenados por ROI Histórico
  - Respeita todos os filtros de rodo (blacklist) do config_universo_97.json
  - Regra de confirmação dupla: Lay_CS_1x0 só entra se Lay_CS_0x1 também aprovado no mesmo jogo
"""

from __future__ import annotations

import io
import json
import zoneinfo
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st

# openpyxl para exportação Excel profissional
try:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── Importar funções do main.py ──────────────────────────────────────────────
import sys
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from main import (
    _load_games_for_date,
    _now_br,
    _compact_source_label,
    _is_local_fallback,
    _is_auth_denied_fallback,
    _is_network_timeout_fallback,
    _render_server_badge,
    _update_connection_state,
)

ROOT_DIR = Path(__file__).resolve().parent.parent
PROD_CFG_PATH = ROOT_DIR / "config_universo_97.json"

# ============================================================================
# CATÁLOGO SUPER ELITE — Métodos com ROI > 10% validado em backtest
# ============================================================================
# Fonte: backtest walk-forward 2023-2026 (config_universo_97.json + engine_ciclo_producao.py)
# ROI calculado sobre stake (responsabilidade), com comissão 6.5%
# EV+ = (prob_green * lucro_green) - (prob_red * stake) > 0
# ============================================================================

SUPER_ELITE_REGISTRY: dict[str, dict[str, Any]] = {
    "Lay_CS_0x1_B365": {
        "status":           "aprovado",
        "roi_historico_pct": 17.5,          # ROI OOS walk-forward 2023-2026
        "prob_green":        0.938,          # taxa de green histórica (93.8%)
        "odd_min":           8.0,
        "odd_max":           11.5,
        "is_lay":            True,
        "descricao":         "LAY Placar 0x1 — Odd 8-11.5",
        "ligas_permitidas":  [               # whitelist do config_universo_97.json
            "AUSTRIA 2","BELGIUM 1","BELGIUM 2","BRAZIL 1","BRAZIL 2","BULGARIA 1",
            "CHILE 1","CHINA 1","COLOMBIA 1","CROATIA 1","CZECH REPUBLIC 1",
            "ECUADOR 1","EGYPT 1","ENGLAND 2","ENGLAND 3","ENGLAND 4","ESTONIA 1",
            "FINLAND 1","FRANCE 2","GREECE 1","HUNGARY 1","IRELAND 1","ISRAEL 1",
            "JAPAN 1","JAPAN 2","MEXICO 1","NORWAY 1","PARAGUAY 1","POLAND 1",
            "PORTUGAL 2","ROMANIA 1","SCOTLAND 1","SERBIA 1","SLOVENIA 1",
            "SOUTH KOREA 1","SWEDEN 1","SWEDEN 2","TURKEY 1","TURKEY 2",
            "USA 1","WALES 1",
        ],
        "blocked_weekdays":  [],
        "prioridade":        1,
        "badge":             "⭐ P1 ELITE",
    },
    "Lay_CS_1x0_B365": {
        "status":           "aprovado",
        "roi_historico_pct": 15.2,          # ROI OOS walk-forward 2023-2026
        "prob_green":        0.879,          # taxa de green histórica (87.9%)
        "odd_min":           4.5,
        "odd_max":           11.5,
        "is_lay":            True,
        "descricao":         "LAY Placar 1x0 — Odd 4.5-11.5 — só SPAIN/ITALY",
        "ligas_permitidas":  ["SPAIN 1", "ITALY 1", "SPAIN 2"],
        "blocked_weekdays":  [],
        "prioridade":        2,
        "badge":             "⭐ P2 ELITE",
        "requer_confirmacao_0x1": True,     # só entra se 0x1 também aprovado no mesmo jogo
    },
}

# Rodos tóxicos (blacklist) — carregados do config
def _carregar_rodos(cfg_path: Path) -> list[dict]:
    try:
        cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
        cuts = []
        seen = set()
        for src in [cfg.get("filtros_rodo", []), cfg.get("filters", {}).get("filtros_rodo", []),
                    cfg.get("filters", {}).get("toxic_cuts", [])]:
            for c in (src or []):
                cid = c.get("id")
                if cid not in seen:
                    seen.add(cid)
                    cuts.append(c)
        return cuts
    except Exception:
        return []


def _is_rodo_bloqueado(liga: str, metodo: str, odd: float, rodos: list[dict]) -> tuple[bool, str]:
    """Retorna (bloqueado, nome_rodo) se o sinal cair em algum rodo tóxico."""
    for cut in rodos:
        cut_leagues = set(cut.get("leagues", []))
        if cut.get("league"):
            cut_leagues.add(str(cut["league"]).upper())
        if cut_leagues and liga.upper() not in cut_leagues:
            continue
        me = cut.get("method_equals")
        mc = cut.get("method_contains")
        if me and str(me) != metodo:
            continue
        if mc and str(mc) not in metodo:
            continue
        omn = cut.get("odd_min")
        omx = cut.get("odd_max")
        if omn is not None and odd < float(omn):
            continue
        if omx is not None and odd > float(omx):
            continue
        return True, str(cut.get("name", f"Rodo_{cut.get('id','?')}"))
    return False, ""


def _calcular_ev(odd: float, prob_green: float, is_lay: bool) -> float:
    """
    EV para Lay: EV = prob_green * (1 - 1/odd) - prob_red * 1
    EV para Back: EV = prob_green * (odd - 1) - prob_red * 1
    Retorna EV por unidade de stake.
    """
    prob_red = 1.0 - prob_green
    if is_lay:
        lucro_green = 1.0 - 1.0 / odd if odd > 1 else 0.0
        return round(prob_green * lucro_green - prob_red * 1.0, 4)
    else:
        return round(prob_green * (odd - 1.0) - prob_red * 1.0, 4)


def _gerar_excel_elite(df_elite: pd.DataFrame, date_str: str) -> bytes:
    """
    Gera Excel profissional com colunas para preenchimento manual:
    Odd_Real_Pega | Stake | Resultado | Lucro_Prejuizo | 1/0
    Layout: cabeçalho azul escuro, bordas, linhas alternadas.
    """
    output = io.BytesIO()

    # Montar DataFrame de exportação
    rows = []
    for _, row in df_elite.iterrows():
        rows.append({
            "Data":           date_str,
            "Hora":           str(row.get("Hora", "") or ""),
            "Jogo":           str(row.get("Jogo", "") or ""),
            "Liga":           str(row.get("Liga", "") or ""),
            "Método":         str(row.get("Método", "") or ""),
            "Odd_Scanner":    round(float(row.get("Odd", 0) or 0), 2),
            "ROI_Hist_%":     row.get("ROI_Hist_%", 0),
            "Green%_Hist":    row.get("Green%_Hist", 0),
            "EV+":            str(row.get("EV_pct", "") or ""),
            # ── Colunas para preenchimento manual ──
            "Odd_Real_Pega":  "",
            "Stake":          "",
            "Resultado":      "",
            "Lucro_Prejuizo": "",
            "1/0":            "",
        })

    df_export = pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
        "Data", "Hora", "Jogo", "Liga", "Método", "Odd_Scanner",
        "ROI_Hist_%", "Green%_Hist", "EV+",
        "Odd_Real_Pega", "Stake", "Resultado", "Lucro_Prejuizo", "1/0",
    ])

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Sinais_Elite")
        ws = writer.sheets["Sinais_Elite"]

        # Estilos
        HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
        HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
        HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
        DATA_FONT    = Font(size=10, name="Calibri")
        CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
        LEFT_ALIGN   = Alignment(horizontal="left", vertical="center")
        THIN_BORDER  = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )
        ALT_FILL     = PatternFill("solid", fgColor="EBF0FA")
        # Colunas de preenchimento manual — fundo amarelo claro
        MANUAL_FILL  = PatternFill("solid", fgColor="FFFACD")

        MANUAL_COLS = {"Odd_Real_Pega", "Stake", "Resultado", "Lucro_Prejuizo", "1/0"}

        ws.row_dimensions[1].height = 28
        for col_idx, col_name in enumerate(df_export.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = HEADER_ALIGN
            cell.border    = THIN_BORDER

        for row_idx in range(2, len(df_export) + 2):
            alt_fill = ALT_FILL if row_idx % 2 == 0 else None
            ws.row_dimensions[row_idx].height = 18
            for col_idx, col_name in enumerate(df_export.columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font   = DATA_FONT
                cell.border = THIN_BORDER
                if col_name in MANUAL_COLS:
                    cell.fill      = MANUAL_FILL
                    cell.alignment = CENTER_ALIGN
                elif col_name in ("Data", "Hora", "Odd_Scanner", "ROI_Hist_%", "Green%_Hist", "EV+"):
                    cell.alignment = CENTER_ALIGN
                    if alt_fill:
                        cell.fill = alt_fill
                else:
                    cell.alignment = LEFT_ALIGN
                    if alt_fill:
                        cell.fill = alt_fill

        # Larguras
        col_widths = {
            "Data": 13, "Hora": 9, "Jogo": 36, "Liga": 22,
            "Método": 22, "Odd_Scanner": 12, "ROI_Hist_%": 12,
            "Green%_Hist": 12, "EV+": 10,
            "Odd_Real_Pega": 14, "Stake": 12,
            "Resultado": 12, "Lucro_Prejuizo": 16, "1/0": 8,
        }
        for col_idx, col_name in enumerate(df_export.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 14)

        ws.freeze_panes = "A2"

        # Aba de instruções
        ws_info = writer.book.create_sheet("Instrucoes")
        ws_info["A1"] = "⭐ ARKAD Super Elite — Planilha de Preenchimento"
        ws_info["A1"].font = Font(bold=True, size=13, color="1F3864", name="Calibri")
        ws_info["A2"] = f"Data: {date_str}  |  Gerado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws_info["A3"] = f"Total de sinais Elite: {len(df_export)}"
        ws_info["A5"] = "COLUNAS PARA PREENCHIMENTO MANUAL (fundo amarelo):"
        ws_info["A5"].font = Font(bold=True, size=11, name="Calibri")
        instrucoes = [
            ("Odd_Real_Pega",  "Odd que você conseguiu executar na Betfair"),
            ("Stake",          "Valor apostado (responsabilidade em R$)"),
            ("Resultado",      "GREEN ou RED"),
            ("Lucro_Prejuizo", "Valor em R$ ganho (+) ou perdido (-)"),
            ("1/0",            "1 = GREEN, 0 = RED"),
        ]
        for i, (col, desc) in enumerate(instrucoes, start=6):
            ws_info[f"A{i}"] = col
            ws_info[f"A{i}"].font = Font(bold=True, name="Calibri")
            ws_info[f"B{i}"] = desc
        ws_info.column_dimensions["A"].width = 20
        ws_info.column_dimensions["B"].width = 50

    output.seek(0)
    return output.read()


def _auditar_sinais_elite(df_jogos: pd.DataFrame, rodos: list[dict]) -> pd.DataFrame:
    """
    Aplica filtro SUPER ELITE em todos os sinais do dia.
    Retorna DataFrame com sinais aprovados + metadados de auditoria.
    """
    if df_jogos.empty:
        return pd.DataFrame()

    resultados = []

    for _, row in df_jogos.iterrows():
        metodo = str(row.get("Metodo", "")).strip()
        cfg_elite = SUPER_ELITE_REGISTRY.get(metodo)

        if cfg_elite is None:
            continue  # método não está no catálogo Elite

        liga = str(row.get("Liga", "")).strip().upper()
        jogo = str(row.get("Jogo", "")).strip()
        hora = str(row.get("Hora", "")).strip()
        odd  = float(row.get("Odd real", 0) or 0)
        status_rodo = str(row.get("Status", "")).strip()

        # ── 1. Status do método ──────────────────────────────────────────────
        if cfg_elite["status"] != "aprovado":
            continue

        # ── 2. ROI > 10% ────────────────────────────────────────────────────
        if cfg_elite["roi_historico_pct"] <= 10.0:
            continue

        # ── 3. Faixa de odd ──────────────────────────────────────────────────
        odd_min = cfg_elite["odd_min"]
        odd_max = cfg_elite["odd_max"]
        if odd < odd_min or odd > odd_max:
            continue

        # ── 4. Liga permitida ────────────────────────────────────────────────
        ligas_perm = {l.upper() for l in cfg_elite.get("ligas_permitidas", [])}
        if ligas_perm and liga not in ligas_perm:
            continue

        # ── 5. Rodo tóxico (blacklist) ───────────────────────────────────────
        bloqueado, nome_rodo = _is_rodo_bloqueado(liga, metodo, odd, rodos)
        if bloqueado:
            continue

        # ── 6. Status do rodo (main.py já aplicou) ───────────────────────────
        if status_rodo == "SKIP":
            continue

        # ── 7. EV+ calculado ─────────────────────────────────────────────────
        prob_green = cfg_elite["prob_green"]
        is_lay = cfg_elite["is_lay"]
        ev = _calcular_ev(odd, prob_green, is_lay)
        if ev <= 0:
            continue  # sem vantagem real → filtrar

        resultados.append({
            "Hora":             hora,
            "Liga":             liga,
            "Jogo":             jogo,
            "Método":           metodo,
            "Badge":            cfg_elite["badge"],
            "Odd":              round(odd, 2),
            "ROI_Hist_%":       cfg_elite["roi_historico_pct"],
            "Green%_Hist":      round(prob_green * 100, 1),
            "EV":               round(ev, 4),
            "EV_pct":           f"{ev*100:+.2f}%",
            "Prioridade":       cfg_elite["prioridade"],
            "Tipo":             "LAY" if is_lay else "BACK",
            "Requer_Conf_0x1":  cfg_elite.get("requer_confirmacao_0x1", False),
            "_jogo_key":        jogo.strip().lower(),
        })

    if not resultados:
        return pd.DataFrame()

    df = pd.DataFrame(resultados)

    # ── 8. Regra de confirmação dupla: 1x0 só entra se 0x1 aprovado no mesmo jogo ──
    jogos_com_0x1 = set(
        df.loc[df["Método"] == "Lay_CS_0x1_B365", "_jogo_key"].tolist()
    )
    mask_1x0_sem_0x1 = (
        (df["Método"] == "Lay_CS_1x0_B365") &
        (df["Requer_Conf_0x1"] == True) &
        (~df["_jogo_key"].isin(jogos_com_0x1))
    )
    df = df[~mask_1x0_sem_0x1].copy()

    # ── 9. Ordenar por ROI Histórico desc, depois por Prioridade e Hora ──────
    df = df.sort_values(
        ["ROI_Hist_%", "Prioridade", "Hora"],
        ascending=[False, True, True]
    ).reset_index(drop=True)

    # ── 10. Limitar a TOP 10 ─────────────────────────────────────────────────
    df = df.head(10)

    # Remover colunas internas
    df = df.drop(columns=["Prioridade", "Requer_Conf_0x1", "_jogo_key"], errors="ignore")

    return df


# ============================================================================
# STREAMLIT UI
# ============================================================================

def main() -> None:
    st.set_page_config(
        page_title="⭐ Auditoria Super Elite — ARKAD",
        layout="wide",
        page_icon="⭐",
    )

    # ── Header ───────────────────────────────────────────────────────────────
    st.markdown("""
    <div style='background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
                padding: 20px 30px; border-radius: 12px; margin-bottom: 20px;
                border: 1px solid #e94560;'>
        <h1 style='color: #FFD700; margin: 0; font-size: 2rem;'>⭐ AUDITORIA SUPER ELITE</h1>
        <p style='color: #aaa; margin: 5px 0 0 0; font-size: 0.95rem;'>
            Apenas sinais com ROI Histórico &gt; 10% + Status Aprovado + EV+ real
            | Máximo TOP 10 por dia | Regra de confirmação dupla ativa
        </p>
    </div>
    """, unsafe_allow_html=True)

    # ── Sidebar ───────────────────────────────────────────────────────────────
    now = _now_br()
    today_iso = now.date().isoformat()
    selected_date = st.sidebar.date_input(
        "📅 Data",
        value=now.date(),
        format="YYYY-MM-DD",
    )
    selected_iso = selected_date.isoformat()
    is_today = selected_iso == today_iso

    st.sidebar.divider()
    st.sidebar.markdown("### ⭐ Critérios SUPER ELITE")
    st.sidebar.markdown("""
    - ✅ ROI Histórico **> 10%**
    - ✅ Status **Aprovado**
    - ✅ **EV+** calculado (vantagem real)
    - ✅ Odd dentro da faixa validada
    - ✅ Liga na whitelist
    - ✅ Sem rodo tóxico (blacklist)
    - ✅ Confirmação dupla (1x0 + 0x1)
    - ✅ **Máximo TOP 10** por dia
    """)

    st.sidebar.divider()
    st.sidebar.markdown("### 📊 Métodos Elite Ativos")
    for nome, cfg in SUPER_ELITE_REGISTRY.items():
        cor = "🟢" if cfg["status"] == "aprovado" else "🔴"
        st.sidebar.markdown(
            f"{cor} **{nome}**  \n"
            f"ROI: +{cfg['roi_historico_pct']}% | Green: {cfg['prob_green']*100:.0f}%  \n"
            f"Odd: {cfg['odd_min']}-{cfg['odd_max']}"
        )

    # ── Carregar dados ────────────────────────────────────────────────────────
    with st.spinner("🔄 Carregando sinais do dia..."):
        df_jogos, source_label = _load_games_for_date(str(PROD_CFG_PATH), selected_iso)
        _update_connection_state(source_label)

    # Status do servidor
    col_srv, col_data = st.columns([2, 1])
    with col_srv:
        _render_server_badge(source_label)
        st.caption(f"Fonte: {_compact_source_label(source_label)}")
    with col_data:
        st.metric("📅 Data Auditada", selected_iso)
        if is_today:
            st.caption(f"🕐 Agora: {now.strftime('%H:%M')} (Brasília)")

    if _is_local_fallback(source_label):
        if _is_auth_denied_fallback(source_label):
            st.error("🔐 Falha de autenticação na API (401/403). Verifique FUTPYTHON_TOKEN.")
        elif _is_network_timeout_fallback(source_label):
            st.info("⏱️ API ao vivo indisponível por timeout. Exibindo base local.")
        else:
            st.warning("⚠️ API indisponível. Exibindo sinais da base local.")

    # ── Carregar rodos ────────────────────────────────────────────────────────
    rodos = _carregar_rodos(PROD_CFG_PATH)

    # ── Auditoria SUPER ELITE ─────────────────────────────────────────────────
    if df_jogos.empty:
        st.info("📭 Sem dados disponíveis para a data selecionada.")
        return

    # Filtrar apenas EXECUTED do main.py
    df_exec = df_jogos[df_jogos["Status"] == "EXECUTED"].copy() if "Status" in df_jogos.columns else df_jogos.copy()

    df_elite = _auditar_sinais_elite(df_exec, rodos)

    # ── Métricas de resumo ────────────────────────────────────────────────────
    total_sinais = len(df_exec)
    total_elite = len(df_elite)
    n_0x1 = len(df_elite[df_elite["Método"] == "Lay_CS_0x1_B365"]) if not df_elite.empty else 0
    n_1x0 = len(df_elite[df_elite["Método"] == "Lay_CS_1x0_B365"]) if not df_elite.empty else 0

    st.divider()
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📋 Sinais do Dia", total_sinais)
    c2.metric("⭐ SUPER ELITE", total_elite, delta=f"de {total_sinais} filtrados")
    c3.metric("🎯 Lay 0x1", n_0x1)
    c4.metric("🎯 Lay 1x0", n_1x0)

    st.divider()

    # ── Tabela SUPER ELITE ────────────────────────────────────────────────────
    if df_elite.empty:
        st.markdown("""
        <div style='background: #1a1a2e; border: 1px solid #e94560; border-radius: 10px;
                    padding: 30px; text-align: center;'>
            <h3 style='color: #e94560;'>🚫 Nenhum sinal SUPER ELITE hoje</h3>
            <p style='color: #aaa;'>Todos os sinais foram filtrados pelos critérios Elite.<br>
            Verifique os rodos tóxicos ou aguarde novos sinais.</p>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown(f"""
        <div style='background: linear-gradient(90deg, #0f3460, #16213e);
                    border: 2px solid #FFD700; border-radius: 10px;
                    padding: 15px 20px; margin-bottom: 15px;'>
            <h3 style='color: #FFD700; margin: 0;'>
                ⭐ TOP {len(df_elite)} SINAIS SUPER ELITE — {selected_iso}
            </h3>
            <p style='color: #aaa; margin: 5px 0 0 0; font-size: 0.85rem;'>
                Ordenados por ROI Histórico (maior primeiro) | Máximo 10 sinais
            </p>
        </div>
        """, unsafe_allow_html=True)

        # Formatar tabela
        df_display = df_elite.copy()

        # Colorir por método
        def _style_row(row):
            if "0x1" in str(row.get("Método", "")):
                return ["background-color: #0d2137; color: white"] * len(row)
            elif "1x0" in str(row.get("Método", "")):
                return ["background-color: #1a0d37; color: white"] * len(row)
            return [""] * len(row)

        # Renomear colunas para exibição
        df_display = df_display.rename(columns={
            "ROI_Hist_%": "ROI Hist.%",
            "Green%_Hist": "Green% Hist.",
            "EV_pct": "EV+",
        })

        cols_show = ["Hora", "Badge", "Liga", "Jogo", "Método", "Odd", "ROI Hist.%", "Green% Hist.", "EV+", "Tipo"]
        cols_show = [c for c in cols_show if c in df_display.columns]

        st.dataframe(
            df_display[cols_show],
            use_container_width=True,
            hide_index=True,
            column_config={
                "Hora":          st.column_config.TextColumn("🕐 Hora", width=70),
                "Badge":         st.column_config.TextColumn("⭐", width=100),
                "Liga":          st.column_config.TextColumn("🏆 Liga", width=150),
                "Jogo":          st.column_config.TextColumn("⚽ Jogo", width=220),
                "Método":        st.column_config.TextColumn("🎯 Método", width=160),
                "Odd":           st.column_config.NumberColumn("📊 Odd", format="%.2f", width=70),
                "ROI Hist.%":    st.column_config.NumberColumn("📈 ROI%", format="+%.1f%%", width=90),
                "Green% Hist.":  st.column_config.NumberColumn("✅ Green%", format="%.1f%%", width=90),
                "EV+":           st.column_config.TextColumn("💡 EV+", width=80),
                "Tipo":          st.column_config.TextColumn("Tipo", width=60),
            },
        )

        # ── Cards individuais ─────────────────────────────────────────────────
        st.markdown("### 📋 Detalhes dos Sinais Elite")
        for i, (_, row) in enumerate(df_elite.iterrows(), 1):
            metodo = str(row.get("Método", ""))
            cor_borda = "#FFD700" if "0x1" in metodo else "#C0C0C0"
            ev_val = float(row.get("EV", 0))
            ev_cor = "#00B050" if ev_val > 0 else "#FF0000"

            st.markdown(f"""
            <div style='background: #0d1b2a; border: 2px solid {cor_borda};
                        border-radius: 10px; padding: 15px 20px; margin-bottom: 10px;'>
                <div style='display: flex; justify-content: space-between; align-items: center;'>
                    <div>
                        <span style='color: #FFD700; font-size: 1.1rem; font-weight: bold;'>
                            #{i} {row.get("Badge","⭐")}
                        </span>
                        <span style='color: #aaa; margin-left: 15px;'>🕐 {row.get("Hora","?")}</span>
                    </div>
                    <div style='text-align: right;'>
                        <span style='color: #00B050; font-weight: bold;'>
                            ROI: +{row.get("ROI_Hist_%", row.get("ROI Hist.%", 0)):.1f}%
                        </span>
                        <span style='color: {ev_cor}; margin-left: 15px; font-weight: bold;'>
                            EV: {ev_val*100:+.2f}%
                        </span>
                    </div>
                </div>
                <div style='margin-top: 8px;'>
                    <span style='color: white; font-size: 1rem;'>
                        ⚽ <strong>{row.get("Jogo","?")}</strong>
                    </span>
                    <span style='color: #aaa; margin-left: 10px;'>
                        🏆 {row.get("Liga","?")}
                    </span>
                </div>
                <div style='margin-top: 5px; color: #aaa; font-size: 0.85rem;'>
                    🎯 {row.get("Método","?")} &nbsp;|&nbsp;
                    📊 Odd: <strong style='color: white;'>{row.get("Odd",0):.2f}</strong> &nbsp;|&nbsp;
                    ✅ Green Hist.: {row.get("Green%_Hist", row.get("Green% Hist.", 0)):.1f}% &nbsp;|&nbsp;
                    {row.get("Tipo","LAY")}
                </div>
            </div>
            """, unsafe_allow_html=True)

    # ── Botão de Download Excel ───────────────────────────────────────────────
    st.divider()
    st.subheader("⬇️ Exportar Sinais para Preenchimento")

    if not df_elite.empty:
        st.caption(
            "Planilha com colunas em **amarelo** para você preencher após os jogos: "
            "**Odd_Real_Pega** · **Stake** · **Resultado** · **Lucro_Prejuizo** · **1/0**"
        )
        col_dl1, col_dl2 = st.columns([3, 1])
        with col_dl2:
            st.metric("Sinais Elite", len(df_elite))
        with col_dl1:
            if OPENPYXL_OK:
                try:
                    xlsx_bytes = _gerar_excel_elite(df_elite, selected_iso)
                    st.download_button(
                        label=f"📥 Baixar Excel — {len(df_elite)} sinais Elite ({selected_iso})",
                        data=xlsx_bytes,
                        file_name=f"elite_{selected_iso}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                    )
                    st.success(
                        "✅ Excel pronto! Colunas **Odd_Real_Pega**, **Stake**, **Resultado**, "
                        "**Lucro_Prejuizo** e **1/0** em branco (fundo amarelo) para preenchimento."
                    )
                except Exception as _e:
                    st.warning(f"⚠️ Erro ao gerar Excel: {_e}")
                    st.download_button(
                        label=f"📥 Baixar CSV — {len(df_elite)} sinais Elite ({selected_iso})",
                        data=df_elite.to_csv(index=False, encoding="utf-8-sig"),
                        file_name=f"elite_{selected_iso}.csv",
                        mime="text/csv",
                        use_container_width=True,
                    )
            else:
                st.download_button(
                    label=f"📥 Baixar CSV — {len(df_elite)} sinais Elite ({selected_iso})",
                    data=df_elite.to_csv(index=False, encoding="utf-8-sig"),
                    file_name=f"elite_{selected_iso}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
    else:
        st.info("Nenhum sinal Elite disponível para exportar.")

    # ── Painel de Governança ──────────────────────────────────────────────────
    st.divider()
    with st.expander("🔒 Governança dos Métodos Elite", expanded=False):
        rows_gov = []
        for nome, cfg in SUPER_ELITE_REGISTRY.items():
            ev_ex = _calcular_ev(9.5, cfg["prob_green"], cfg["is_lay"])
            rows_gov.append({
                "Método":          nome,
                "Status":          cfg["status"].upper(),
                "ROI Hist.%":      f"+{cfg['roi_historico_pct']}%",
                "Green% Hist.":    f"{cfg['prob_green']*100:.1f}%",
                "Odd Min":         cfg["odd_min"],
                "Odd Max":         cfg["odd_max"],
                "EV (odd=9.5)":    f"{ev_ex*100:+.2f}%",
                "Ligas":           len(cfg.get("ligas_permitidas", [])),
                "Tipo":            "LAY" if cfg["is_lay"] else "BACK",
            })
        st.dataframe(pd.DataFrame(rows_gov), use_container_width=True, hide_index=True)

        st.markdown(f"""
        **Rodos Tóxicos Ativos:** {len(rodos)} combinações bloqueadas  
        **Regra de Confirmação Dupla:** Lay_CS_1x0 só entra se Lay_CS_0x1 aprovado no mesmo jogo  
        **Trava EV+:** Sinais sem vantagem real calculada são filtrados automaticamente  
        **Cap Diário:** Máximo 10 sinais exibidos (ordenados por ROI Histórico)
        """)

    # ── Todos os sinais do dia (para referência) ──────────────────────────────
    with st.expander(f"📋 Todos os Sinais Aprovados do Dia ({len(df_exec)})", expanded=False):
        if df_exec.empty:
            st.info("Sem sinais aprovados.")
        else:
            cols_all = [c for c in ["Hora", "Prio", "Liga", "Jogo", "Metodo", "Odd real", "Status"] if c in df_exec.columns]
            st.dataframe(df_exec[cols_all], use_container_width=True, hide_index=True)

    # ── Rodapé ────────────────────────────────────────────────────────────────
    st.divider()
    st.caption(
        f"⭐ ARKAD Super Elite | Auditoria: {selected_iso} | "
        f"Gerado: {now.strftime('%d/%m/%Y %H:%M')} (Brasília) | "
        f"Métodos ativos: {len([m for m in SUPER_ELITE_REGISTRY.values() if m['status']=='aprovado'])}"
    )


if __name__ == "__main__":
    main()
